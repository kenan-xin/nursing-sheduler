"""Focused tests for exporter helpers and formatting edge cases."""

# This file is part of Nurse Scheduling Project, see <https://github.com/j3soon/nurse-scheduling>.
#
# Copyright (C) 2023-2026 Johnson Sun
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

# This test is mostly AI generated.

import os
import sys
from io import BytesIO
from types import SimpleNamespace

import pandas as pd
import pytest
from openpyxl import load_workbook

# Add the project root to the Python path so imports will work when running directly
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from nurse_scheduling import exporter, schedule


@pytest.mark.parametrize(
    ("shift_types", "values", "weight", "expected"),
    [
        ([0, 1], {"D": 1, "E": 0, "OFF": 0}, 5, True),
        ([0, 1], {"D": 1, "E": 0, "OFF": 0}, -5, False),
        ([0, 1], {"D": 0, "E": 0, "OFF": 1}, -5, True),
        ([exporter.constants.OFF_sid], {"D": 0, "E": 0, "OFF": 1}, 5, True),
        ([exporter.constants.OFF_sid], {"D": 1, "E": 0, "OFF": 0}, -5, True),
    ],
)
def test_shift_request_satisfaction_uses_any_requested_state_then_weight_sign(shift_types, values, weight, expected):
    ctx = SimpleNamespace(
        solver=SimpleNamespace(get_value=lambda var: values[var]),
        shifts={(0, 0, 0): "D", (0, 1, 0): "E"},
        offs={(0, 0): "OFF"},
    )
    pref = SimpleNamespace(weight=weight)

    assert exporter._is_shift_request_satisfied(ctx, pref, d=0, p=0, shift_types=shift_types) is expected


def test_export_to_csv_writes_utf8_bom():
    df = pd.DataFrame([["A", "B"], ["C", "D"]])
    output = BytesIO()

    exporter.export_to_csv(df, output)

    payload = output.getvalue()
    assert payload.startswith(b"\xef\xbb\xbf")
    assert payload.decode("utf-8-sig") == "A,B\nC,D\n"


def test_export_to_excel_rejects_legacy_comment_info_shape():
    df = pd.DataFrame([["x"]])
    output = BytesIO()

    with pytest.raises(ValueError, match="cell_export_info must be a dictionary"):
        exporter.export_to_excel(df, output, {(1, 1): [3, 7]})


def test_export_to_excel_applies_style_and_font_contrast():
    df = pd.DataFrame([["dark", "light"]])
    output = BytesIO()
    cell_export_info = {
        "comments": {},
        "styles": {
            (1, 1): {"backgroundColor": "#111111"},
            (1, 2): {
                "backgroundColor": "#f5f5f5",
                "bottomBorderColor": "#0ea5e9",
                "rightBorderColor": "#9ca3af",
                "fontColor": "#dc2626",
            },
        },
    }

    exporter.export_to_excel(df, output, cell_export_info)

    wb = load_workbook(output)
    ws = wb.active
    assert wb.sheetnames == ["Sheet1"]
    assert ws["A1"].fill.fgColor.rgb == "FF111111"
    assert ws["A1"].font.color is not None
    assert ws["A1"].font.color.rgb == "FFFFFFFF"
    assert ws["B1"].fill.fgColor.rgb == "FFF5F5F5"
    assert ws["B1"].font.color is not None
    assert ws["B1"].font.color.rgb == "FFDC2626"
    assert ws["B1"].border.bottom.color is not None
    assert ws["B1"].border.bottom.color.rgb == "FF0EA5E9"
    assert ws["B1"].border.bottom.style == "medium"
    assert ws["B1"].border.right.color is not None
    assert ws["B1"].border.right.color.rgb == "FF9CA3AF"
    assert ws["B1"].border.right.style == "medium"


def test_prettify_exports_notes_sheet_for_unmet_single_style_requests():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-02
people:
  items:
    - id: n1
      history: [D]
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 1
    qualifiedPeople: ALL
    date: ALL
    weight: -1
  - type: shift request
    person: n1
    date: ["2025-01-01"]
    shiftType: D
    weight: -10
export:
  formatting:
    - type: cell
      people: [ALL]
      dates: [ALL]
      shiftTypes: [ALL, OFF]
      when:
        preference:
          types: ["shift request"]
          requestShape: [person-item-to-date-item]
          satisfied: false
          weightRange: [-.inf, .inf]
      appendText: " [X]"
      note:
        text: "Weight of unmet single-style request: {totalAbsWeight}"
"""

    df, _solution, _score, _status, cell_export_info = schedule(yaml_content, prettify=True)
    assert hasattr(df, "to_excel")
    assert "comments" in cell_export_info
    assert "styles" in cell_export_info
    assert cell_export_info["comments"]

    first_target_cell = next(iter(cell_export_info["comments"].keys()))
    output = BytesIO()
    exporter.export_to_excel(df, output, cell_export_info)

    wb = load_workbook(output)
    ws = wb.active
    row, col = first_target_cell
    schedule_cell = ws.cell(row=row, column=col)
    notes_ws = wb["Notes"]
    assert schedule_cell.comment is None
    assert schedule_cell.hyperlink is not None
    assert schedule_cell.hyperlink.target == "#'Notes'!A2"
    assert schedule_cell.fill.fill_type is None
    assert notes_ws.freeze_panes == "A2"
    assert notes_ws.auto_filter.ref == "A1:C1"
    assert [cell.value for cell in notes_ws[1]] == ["Cell", "Schedule Value", "Note"]
    assert notes_ws["A2"].value == schedule_cell.coordinate
    assert notes_ws["A2"].hyperlink is not None
    assert notes_ws["A2"].hyperlink.target == f"#'{ws.title}'!{schedule_cell.coordinate}"
    assert notes_ws["C2"].value == "Weight of unmet single-style request: 10"


def test_export_to_excel_writes_each_note_as_a_separate_notes_sheet_row():
    df = pd.DataFrame([["annotated"]])
    output = BytesIO()

    exporter.export_to_excel(df, output, {"comments": {(1, 1): ["first", "second"]}, "styles": {}})

    wb = load_workbook(output)
    ws = wb.active
    notes_ws = wb["Notes"]
    assert ws["A1"].comment is None
    assert ws["A1"].hyperlink is not None
    assert ws["A1"].hyperlink.target == "#'Notes'!A2"
    assert ws["A1"].fill.fill_type is None
    assert list(notes_ws.values) == [
        ("Cell", "Schedule Value", "Note"),
        ("A1", "annotated", "first"),
        ("A1", "annotated", "second"),
    ]
    assert notes_ws["A2"].hyperlink is not None
    assert notes_ws["A2"].hyperlink.target == "#'Sheet1'!A1"
    assert notes_ws["A3"].hyperlink is not None
    assert notes_ws["A3"].hyperlink.target == "#'Sheet1'!A1"


def test_export_annotations_expand_compacted_shift_request_dates_before_matching_shape():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-02
  groups:
    - id: FREEDAY
      members: [2025-01-02]
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 0
  - type: shift request
    person: n1
    date: ["01", FREEDAY]
    shiftType: D
    weight: -5
export:
  formatting:
    - type: cell
      appendText: " [I]"
      people: [ALL]
      dates: [ALL]
      shiftTypes: [D]
      when:
        preference:
          types: ["shift request"]
          requestShape: [person-item-to-date-item]
    - type: cell
      appendText: " [G]"
      people: [ALL]
      dates: [ALL]
      shiftTypes: [D]
      when:
        preference:
          types: ["shift request"]
          requestShape: [person-item-to-date-group]
"""

    styled_df, _solution, _score, _status, _cell_export_info = schedule(yaml_content, prettify=True)
    df = styled_df.data

    assert str(df.iloc[2, 1]) == " [I]"
    assert str(df.iloc[2, 2]) == " [G]"


def test_export_annotation_total_abs_weight_sums_matched_requests_for_cell():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-01
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
    - id: E
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: [D, E]
    requiredNumPeople: 0
  - type: shift request
    person: n1
    date: ["2025-01-01"]
    shiftType: D
    weight: -5
  - type: shift request
    person: n1
    date: ["2025-01-01"]
    shiftType: E
    weight: -7
export:
  formatting:
    - type: cell
      note:
        text: "Total matched absolute weight: {totalAbsWeight}"
      people: [ALL]
      dates: [ALL]
      shiftTypes: [ALL]
      when:
        preference:
          types: ["shift request"]
          requestShape: [person-item-to-date-item]
          satisfied: true
"""

    _df, _solution, _score, _status, cell_export_info = schedule(yaml_content, prettify=True)

    assert cell_export_info["comments"] == {(3, 2): ["Total matched absolute weight: 12"]}


def test_export_annotation_treats_multi_shift_type_request_shape_as_unknown():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-01
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
    - id: E
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: [D, E]
    requiredNumPeople: 0
  - type: shift request
    person: n1
    date: ["2025-01-01"]
    shiftType: [D, E]
    weight: 5
export:
  formatting:
    - type: cell
      appendText: " [specific]"
      people: [ALL]
      dates: [ALL]
      shiftTypes: [ALL]
      when:
        preference:
          types: ["shift request"]
          requestShape: [person-item-to-date-item]
          weightRange: [-.inf, .inf]
    - type: cell
      appendText: " [all:{shiftType}]"
      note:
        text: "Total matched absolute weight: {totalAbsWeight}"
      people: [ALL]
      dates: [ALL]
      shiftTypes: [ALL]
      when:
        preference:
          types: ["shift request"]
          requestShape: [ALL]
          weightRange: [-.inf, .inf]
"""

    styled_df, _solution, _score, _status, cell_export_info = schedule(yaml_content, prettify=True)

    assert str(styled_df.data.iloc[2, 1]) == " [all:D, E]"
    assert cell_export_info["comments"] == {(3, 2): ["Total matched absolute weight: 5"]}


def test_export_annotation_marks_assigned_negative_shift_type_group_as_unsatisfied():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-01
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
    - id: E
  groups:
    - id: DayOrEvening
      members: [D, E]
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 1
  - type: shift type requirement
    shiftType: E
    requiredNumPeople: 0
  - type: shift request
    person: n1
    date: ["2025-01-01"]
    shiftType: DayOrEvening
    weight: -5
export:
  formatting:
    - type: cell
      appendText: " [unsatisfied]"
      people: [ALL]
      dates: [ALL]
      shiftTypes: [ALL]
      when:
        preference:
          types: ["shift request"]
          satisfied: false
"""

    styled_df, _solution, _score, _status, _cell_export_info = schedule(yaml_content, prettify=True)

    assert str(styled_df.data.iloc[2, 1]) == "D [unsatisfied]"


def test_export_annotation_rejects_reversed_weight_range():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-01
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 0
  - type: shift request
    person: n1
    date: ["2025-01-01"]
    shiftType: D
    weight: -5
export:
  formatting:
    - type: cell
      appendText: " [X]"
      people: [ALL]
      dates: [ALL]
      shiftTypes: [D]
      when:
        preference:
          types: ["shift request"]
          weightRange: [10, -10]
"""

    with pytest.raises(ValueError, match="weightRange minimum"):
        schedule(yaml_content, prettify=True)


def test_export_formatting_rejects_non_cell_when_and_annotations():
    base_ctx = SimpleNamespace(
        export=SimpleNamespace(formatting=[]),
        map_pid_p={"n1": [0]},
        map_did_d={},
        map_sid_s={},
    )

    base_ctx.export.formatting = [
        SimpleNamespace(
            type="row",
            people=["n1"],
            backgroundColor="#22c55e",
            bottomBorderColor=None,
            rightBorderColor=None,
            fontColor=None,
            when=SimpleNamespace(),
            appendText=None,
            note=None,
        )
    ]
    with pytest.raises(ValueError, match="'when' is only supported"):
        exporter._build_custom_export_style_info(
            base_ctx,
            n_rows=1,
            n_cols=1,
            n_leading_rows=0,
            n_leading_cols=0,
            n_history_cols=0,
        )

    base_ctx.export.formatting = [
        SimpleNamespace(
            type="row",
            people=["n1"],
            backgroundColor="#22c55e",
            bottomBorderColor=None,
            rightBorderColor=None,
            fontColor=None,
            when=None,
            appendText=" [X]",
            note=None,
        )
    ]
    with pytest.raises(ValueError, match="annotations are only supported"):
        exporter._build_custom_export_style_info(
            base_ctx,
            n_rows=1,
            n_cols=1,
            n_leading_rows=0,
            n_leading_cols=0,
            n_history_cols=0,
        )


def test_export_annotation_unknown_request_shape_matches_all_but_not_specific_shape():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-01
people:
  items:
    - id: n1
    - id: n2
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 0
  - type: shift request
    person: [n1, n2]
    date: ["2025-01-01"]
    shiftType: D
    weight: -5
export:
  formatting:
    - type: cell
      appendText: " [specific]"
      people: [ALL]
      dates: [ALL]
      shiftTypes: [D]
      when:
        preference:
          types: ["shift request"]
          requestShape: [person-item-to-date-item]
    - type: cell
      appendText: " [all]"
      people: [ALL]
      dates: [ALL]
      shiftTypes: [D]
      when:
        preference:
          types: ["shift request"]
          requestShape: [ALL]
"""

    styled_df, _solution, _score, _status, _cell_export_info = schedule(yaml_content, prettify=True)
    df = styled_df.data

    assert str(df.iloc[2, 1]) == " [all]"
    assert str(df.iloc[3, 1]) == " [all]"


def test_invalid_row_target_in_export_formatting_raises():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-01
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 1
    qualifiedPeople: ALL
    date: ALL
    weight: -1
export:
  formatting:
    - type: row
      people: [unknown_person]
      backgroundColor: "#22c55e"
"""

    with pytest.raises(ValueError, match="Invalid person identifier"):
        schedule(yaml_content, prettify=False)


def test_invalid_cell_target_in_export_formatting_raises():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-01
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 1
    qualifiedPeople: ALL
    date: ALL
    weight: -1
export:
  formatting:
    - type: cell
      people: [ALL]
      dates: [ALL]
      shiftTypes: [UNKNOWN_SHIFT]
      backgroundColor: "#ef4444"
"""

    with pytest.raises(ValueError, match="Invalid shift type identifier"):
        schedule(yaml_content, prettify=False)


def test_date_headers_format_with_year_boundary():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2024-12-31
    endDate: 2025-01-02
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 0
"""

    df, _solution, _score, _status, _cell_export_info = schedule(yaml_content, prettify=False)
    assert df.iloc[0, 1] == "2024/12/31"
    assert df.iloc[0, 2] == "2025/1/1"
    assert df.iloc[0, 3] == "2025/1/2"


def test_prettify_off_annotations_and_workday_freeday_headers():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-03
  groups:
    - id: WORKDAY
      members: [2025-01-02, 2025-01-03]
    - id: FREEDAY
      members: [2025-01-01]
people:
  items:
    - id: n1
      history: [D]
    - id: n2
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 0
  - type: shift request
    person: n1
    date: ["2025-01-01"]
    shiftType: [D]
    weight: 0
  - type: shift request
    person: n1
    date: ["2025-01-01"]
    shiftType: [OFF]
    weight: -5
export:
  formatting:
    - type: cell
      people: [ALL]
      dates: [ALL]
      shiftTypes: [ALL, OFF]
      when:
        preference:
          types: ["shift request"]
          requestShape: [person-item-to-date-item]
          weightRange: [-.inf, .inf]
      appendText: " [{shiftType}]"
  extraColumns:
    - type: count
      header: OFF (WORKDAY)
      countShiftTypes: [OFF]
      countDates: [WORKDAY]
    - type: count
      header: OFF (FREEDAY)
      countShiftTypes: [OFF]
      countDates: [FREEDAY]
  extraRows:
    - type: count
      header: OFF Count
      countShiftTypes: [OFF]
      countPeople: [ALL]
"""
    styled_df, _solution, _score, _status, _cell_export_info = schedule(yaml_content, prettify=True)
    df = styled_df.data

    # Weight-0 shift request should be ignored by prettify markers,
    # while OFF request should still annotate the cell.
    target_cell = str(df.iloc[2, 2])
    assert "[OFF]" in target_cell
    assert "[D]" not in target_cell

    # History fallback branch for person without history.
    assert df.iloc[3, 1] == ""

    # Workday/freeday summary headers should be present when both groups are found.
    headers = list(df.iloc[1, :])
    assert "OFF (WORKDAY)" in headers
    assert "OFF (FREEDAY)" in headers
    assert df.iloc[7, 0] == "OFF Count"
    assert df.iloc[7, 2] == 2


def test_export_extra_column_counts_shift_type_coefficient_scores():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-02
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
    - id: A
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 1
    date: 2025-01-01
  - type: shift type requirement
    shiftType: A
    requiredNumPeople: 1
    date: 2025-01-02
export:
  extraColumns:
    - type: count
      header: Weighted Score
      countShiftTypes: [D, A]
      countShiftTypeCoefficients:
        - [D, 2]
        - [A, 3]
      countDates: [ALL]
"""
    styled_df, _solution, _score, _status, cell_export_info = schedule(yaml_content, prettify=True)
    df = styled_df.data

    assert df.iloc[1, 4] == "Weighted Score"
    assert df.iloc[2, 4] == 5

    output = BytesIO()
    exporter.export_to_excel(styled_df, output, cell_export_info)
    workbook = load_workbook(output)
    assert workbook.active.cell(row=3, column=5).value == 5


def test_export_extra_column_defaults_to_one_and_scores_off_coefficients():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-02
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 1
    date: 2025-01-01
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 0
    date: 2025-01-02
export:
  extraColumns:
    - type: count
      header: Default Count
      countShiftTypes: [D, OFF]
      countDates: [ALL]
    - type: count
      header: Weighted Count
      countShiftTypes: [D, OFF]
      countShiftTypeCoefficients:
        - [D, 2]
        - [OFF, 4]
      countDates: [ALL]
"""
    styled_df, _solution, _score, _status, _cell_export_info = schedule(yaml_content, prettify=True)
    df = styled_df.data

    assert df.iloc[2, 4] == 2
    assert df.iloc[2, 5] == 6


def test_export_extra_column_applies_group_coefficient_to_expanded_shift_type():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-01
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
    - id: A
  groups:
    - id: WORK
      members: [D, A]
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 1
export:
  extraColumns:
    - type: count
      header: Work Score
      countShiftTypes: [WORK]
      countShiftTypeCoefficients:
        - [WORK, 7]
      countDates: [ALL]
"""
    styled_df, _solution, _score, _status, _cell_export_info = schedule(yaml_content, prettify=True)
    df = styled_df.data

    assert df.iloc[2, 3] == 7


def test_export_extra_column_applies_member_coefficient_covered_by_selected_group():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-01
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
    - id: A
  groups:
    - id: WORK
      members: [D, A]
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 1
export:
  extraColumns:
    - type: count
      header: Work Score
      countShiftTypes: [WORK]
      countShiftTypeCoefficients:
        - [D, 7]
      countDates: [ALL]
"""
    styled_df, _solution, _score, _status, _cell_export_info = schedule(yaml_content, prettify=True)
    df = styled_df.data

    assert df.iloc[2, 3] == 7


def test_build_custom_export_style_info_ignores_out_of_bounds_targets():
    ctx = SimpleNamespace(
        export=SimpleNamespace(
            formatting=[
                SimpleNamespace(
                    type="row",
                    people=["n1"],
                    backgroundColor="#22c55e",
                    bottomBorderColor=None,
                    rightBorderColor=None,
                    fontColor=None,
                )
            ]
        ),
        map_pid_p={"n1": [0]},
        map_did_d={},
        map_sid_s={},
    )

    # n_rows=0 forces set_style to hit out-of-bounds guard and skip writes.
    style_map = exporter._build_custom_export_style_info(
        ctx,
        n_rows=0,
        n_cols=1,
        n_leading_rows=2,
        n_leading_cols=1,
        n_history_cols=0,
    )
    assert style_map == {}


def test_dataframe_generation_supports_multiple_assigned_shift_types():
    class DummySolver:
        def get_value(self, var):
            return 1 if var in {"v_d", "v_e"} else 0

        def get_objective_value(self):
            return 0

    ctx = SimpleNamespace(
        n_shift_types=2,
        shiftTypes=SimpleNamespace(
            items=[SimpleNamespace(id="D"), SimpleNamespace(id="E")],
            groups=[],
        ),
        people=SimpleNamespace(items=[SimpleNamespace(id="n1", history=None)]),
        dates=SimpleNamespace(
            items=[
                SimpleNamespace(
                    year=2025, month=1, day=1, weekday=lambda: 2, strftime=lambda fmt: "Wed" if fmt == "%a" else "1"
                )
            ],
            groups=[],
            range=SimpleNamespace(
                startDate=SimpleNamespace(year=2025, month=1), endDate=SimpleNamespace(year=2025, month=1)
            ),
        ),
        map_dp_s={(0, 0): {0, 1}},
        shifts={(0, 0, 0): "v_d", (0, 1, 0): "v_e"},
        offs={(0, 0): "v_off"},
        preferences=[],
        map_sid_s={},
        map_pid_p={},
        map_did_d={},
        solver=DummySolver(),
        solver_status="OPTIMAL",
        export=None,
    )

    df, info = exporter.get_people_versus_date_dataframe(ctx, prettify=False)
    assert df.iloc[2, 1] == "D, E"
    assert info["styles"] == {}


def test_prettify_styling_does_not_add_default_freeday_or_weekend_colors():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-03
    endDate: 2025-01-05
  groups:
    - id: WORKDAY
      members: [2025-01-03]
    - id: FREEDAY
      members: [2025-01-05]
people:
  items:
    - id: n1
      history: [D]
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 0
  - type: shift request
    person: n1
    date: ["2025-01-03"]
    shiftType: [OFF]
    weight: -5
"""
    styled_df, _solution, _score, _status, _cell_export_info = schedule(yaml_content, prettify=True)
    html = styled_df.to_html()
    assert "text-align: center" in html
    assert "background-color: #fefce8" not in html
    assert "#dcfce7" not in html
    assert "#dbeafe" not in html
    assert "#9ca3af" not in html
