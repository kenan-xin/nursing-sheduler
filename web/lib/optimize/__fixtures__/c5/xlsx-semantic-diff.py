#!/usr/bin/env python3
# Independent openpyxl semantic diff for the T16c restoration tests.
#
# Reads two workbook paths from argv (original, restored) and prints, as JSON on
# stdout, EVERY semantic difference openpyxl can see between them: sheet set,
# per-sheet dimensions/freeze panes, and per-cell value, data type, number
# format, font, fill, border, alignment, and hyperlink. The Vitest suite asserts
# the only reported differences are the intended column-A people-id cells — a
# check that is genuinely independent of the TypeScript surgical editor (a
# different library, a different parse).
#
# Usage:  python xlsx-semantic-diff.py <original.xlsx> <restored.xlsx>
# Exits non-zero only on an internal error; a clean or dirty diff both exit 0.

import json
import sys

from openpyxl import load_workbook


def cell_style_signature(cell):
    font = cell.font
    fill = cell.fill
    border = cell.border
    align = cell.alignment
    return {
        "number_format": cell.number_format,
        "font": (font.name, font.size, font.bold, font.italic, font.underline,
                 font.color.rgb if font.color else None),
        "fill": (fill.fill_type, getattr(fill.fgColor, "rgb", None),
                 getattr(fill.bgColor, "rgb", None)),
        "border": tuple(
            (getattr(side, "style", None), getattr(getattr(side, "color", None), "rgb", None))
            for side in (border.left, border.right, border.top, border.bottom)
        ),
        "align": (align.horizontal, align.vertical, align.wrap_text),
        "hyperlink": cell.hyperlink.target if cell.hyperlink else None,
    }


def sheet_cells(ws):
    cells = {}
    for row in ws.iter_rows():
        for cell in row:
            cells[cell.coordinate] = cell
    return cells


def diff(orig_path, restored_path):
    wb_a = load_workbook(orig_path)
    wb_b = load_workbook(restored_path)

    diffs = []
    if wb_a.sheetnames != wb_b.sheetnames:
        diffs.append({"kind": "sheetnames", "a": wb_a.sheetnames, "b": wb_b.sheetnames})
        return diffs

    for name in wb_a.sheetnames:
        ws_a, ws_b = wb_a[name], wb_b[name]
        if (ws_a.max_row, ws_a.max_column) != (ws_b.max_row, ws_b.max_column):
            diffs.append({"kind": "dimensions", "sheet": name,
                          "a": [ws_a.max_row, ws_a.max_column],
                          "b": [ws_b.max_row, ws_b.max_column]})
        if ws_a.freeze_panes != ws_b.freeze_panes:
            diffs.append({"kind": "freeze_panes", "sheet": name,
                          "a": ws_a.freeze_panes, "b": ws_b.freeze_panes})

        cells_a, cells_b = sheet_cells(ws_a), sheet_cells(ws_b)
        for coord in sorted(set(cells_a) | set(cells_b)):
            ca, cb = cells_a.get(coord), cells_b.get(coord)
            if ca is None or cb is None:
                diffs.append({"kind": "presence", "sheet": name, "coord": coord})
                continue
            if ca.value != cb.value:
                diffs.append({"kind": "value", "sheet": name, "coord": coord,
                              "a": _jsonable(ca.value), "b": _jsonable(cb.value),
                              "a_type": ca.data_type, "b_type": cb.data_type})
            if cell_style_signature(ca) != cell_style_signature(cb):
                diffs.append({"kind": "style", "sheet": name, "coord": coord})
    return diffs


def _jsonable(value):
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def main():
    if len(sys.argv) != 3:
        print(json.dumps({"error": "usage: xlsx-semantic-diff.py <a> <b>"}))
        sys.exit(2)
    print(json.dumps({"diffs": diff(sys.argv[1], sys.argv[2])}))


if __name__ == "__main__":
    main()
