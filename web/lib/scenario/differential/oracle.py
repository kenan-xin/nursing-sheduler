"""Differential oracle (T05) — the real backend the C1/C3/C5 harnesses run against.

This is NOT a reimplementation of anything: it imports the *vendored* Python
backend (`core/nurse_scheduling`) and drives its actual `load_data`, `schedule`,
`group_map`, and `exporter` code so the TypeScript contract layer is checked
against binding behavior rather than a memory of it.

Protocol: read one JSON request object from stdin, write one JSON response object
to stdout. The process exits 0 even for scenario-level rejections (they are the
data); a non-zero exit means the harness itself broke. Ops:

  {"op": "load",       "yaml": "..."}                 -> C1: load_data + model dump
  {"op": "schedule",   "yaml": "..."}                 -> C3: full schedule accept/reject
  {"op": "shift_map",  "items": [...], "groups": [...]}-> C3: ordered group-map construction
  {"op": "export",     "yaml": "..."}                 -> C5: real exporter (df + xlsx)
  {"op": "roundtrip",  "raw": "...", "roundtrip": "..."} -> import round-trip equivalence
"""

import io
import json
import os
import sys

# Import the vendored backend regardless of cwd: this file lives at
# <root>/web/lib/scenario/differential/oracle.py; the backend is at <root>/core.
_HERE = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.abspath(os.path.join(_HERE, "..", "..", "..", ".."))
sys.path.insert(0, os.path.join(_ROOT, "core"))

import nurse_scheduling  # noqa: E402
from nurse_scheduling import exporter, group_map  # noqa: E402
from nurse_scheduling.constants import ALL, MAP_DATE_KEYWORD_TO_FILTER, MAP_WEEKDAY_TO_STR  # noqa: E402
from nurse_scheduling.loader import load_data  # noqa: E402
from nurse_scheduling.models import DateRange, ShiftType, ShiftTypeGroup  # noqa: E402
from nurse_scheduling.server.scheduling_errors import SchedulingContentError  # noqa: E402
from nurse_scheduling.utils import parse_dates, parse_pids  # noqa: E402
from nurse_scheduling.server.scheduling_input import (  # noqa: E402
    MalformedInputError,
    canonicalize_submission,
)
from openpyxl import load_workbook  # noqa: E402


def _dump(model):
    return model.model_dump(mode="json")


def _json_safe(value):
    """Recursively convert non-finite floats to sentinel strings so the output is
    valid JSON (JS `JSON.parse` rejects bare `Infinity`/`NaN`). The mapping is
    stable, so round-trip comparisons over sanitized dumps stay meaningful."""
    if isinstance(value, float):
        if value == float("inf"):
            return "Infinity"
        if value == float("-inf"):
            return "-Infinity"
        if value != value:  # NaN
            return "NaN"
        return value
    if isinstance(value, dict):
        return {k: _json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    return value


def op_load(req):
    """C1: bytes -> load_data. Accept -> canonical model dump; reject -> error."""
    try:
        data = load_data(req["yaml"].encode("utf-8"))
        return {"ok": True, "model": _dump(data)}
    except Exception as e:  # noqa: BLE001 — every rejection is data here
        return {"ok": False, "error": str(e), "errorType": type(e).__name__}


def op_schedule(req):
    """C3: run the real scheduler/context setup. Rejections raised during model
    build (OFF/LEAVE misuse, unknown ids, coverage) surface as {ok: false}."""
    try:
        _df, _sol, _score, status, _cell = nurse_scheduling.schedule(req["yaml"].encode("utf-8"))
        return {"ok": True, "status": status}
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "error": str(e), "errorType": type(e).__name__}


def op_shift_map(req):
    """C3: the ordered shift-type id -> [indices] map (group_map). Verifies
    ALL/OFF/LEAVE expansion, definition-order construction, and forward-ref failure."""
    try:
        items = [ShiftType(id=i) for i in req["items"]]
        groups = [ShiftTypeGroup(id=g["id"], members=g["members"]) for g in req["groups"]]
        result = group_map.build_shift_type_index_map(items, groups)
        return {"ok": True, "map": {str(k): v for k, v in result.items()}}
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "error": str(e), "errorType": type(e).__name__}


def _typed_key_records(mapping):
    """Emit one lossless record per map entry in insertion order, preserving the
    original key TYPE and value. This is the transport that distinguishes numeric
    `1` from string `"1"` — a JSON object keyed by `str(k)` would collide them."""
    records = []
    for key, indices in mapping.items():
        records.append(
            {
                "keyType": "number" if isinstance(key, (int, float)) and not isinstance(key, bool) else "string",
                "key": key,
                "indices": list(indices),
            }
        )
    return records


def _build_people_index_map(staff, groups):
    """Mirror scheduler.py `ctx.map_pid_p` construction: raw typed person ids, then
    the `ALL` keyword, then people groups in declaration order resolving through
    the map built so far. Keys keep their original int/str identity."""
    map_pid_p = {}
    n_people = len(staff)
    for p in range(n_people):
        map_pid_p[staff[p]] = [p]
    map_pid_p[ALL] = list(range(n_people))
    for group in groups:
        map_pid_p[group["id"]] = sorted(set().union(*[map_pid_p[pid] for pid in group["members"]]) if group["members"] else set())
    return map_pid_p


def _build_date_index_map(date_range, groups):
    """Mirror scheduler.py `ctx.map_did_d` construction: each ISO date, then the
    date-filter keywords, then weekday names, then date groups (direct-lookup-then-
    parse per member) in declaration order."""
    import datetime

    startdate, enddate = date_range.startDate, date_range.endDate
    n_days = (enddate - startdate).days + 1
    items = [startdate + datetime.timedelta(days=d) for d in range(n_days)]
    map_did_d = {}
    for d in range(n_days):
        map_did_d[str(items[d])] = [d]
    for keyword in MAP_DATE_KEYWORD_TO_FILTER:
        map_did_d[keyword] = [d for d in range(n_days) if MAP_DATE_KEYWORD_TO_FILTER[keyword](items[d])]
    for keyword in MAP_WEEKDAY_TO_STR:
        weekday_index = MAP_WEEKDAY_TO_STR.index(keyword)
        map_did_d[keyword] = [d for d in range(n_days) if items[d].weekday() == weekday_index]
    for group in groups:
        date_indices = set()
        for member in group["members"]:
            if member in map_did_d:
                date_indices.update(map_did_d[member])
            else:
                date_indices.update(parse_dates(member, map_did_d, date_range))
        map_did_d[group["id"]] = sorted(set(date_indices))
    return map_did_d


def op_people_map(req):
    """C3: the ordered person-id -> [indices] map as LOSSLESS tagged records, so a
    numeric/string person-id collision is observable rather than merged."""
    try:
        map_pid_p = _build_people_index_map(req["staff"], req.get("groups", []))
        return {"ok": True, "records": _typed_key_records(map_pid_p)}
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "error": str(e), "errorType": type(e).__name__}


def op_resolve_people(req):
    """C3: resolve a person selector through `parse_pids` over the ordered people
    map. Returns the discriminated {resolved, indices}; an unknown id is unresolved."""
    try:
        map_pid_p = _build_people_index_map(req["staff"], req.get("groups", []))
    except Exception as e:  # noqa: BLE001 — map construction failure is data
        return {"ok": True, "resolved": False, "error": str(e)}
    try:
        indices = parse_pids(req["selector"], map_pid_p)
        return {"ok": True, "resolved": True, "indices": indices}
    except Exception as e:  # noqa: BLE001
        return {"ok": True, "resolved": False, "error": str(e)}


def op_resolve_dates(req):
    """C3: resolve a date selector through `parse_dates` over the ordered date map.
    The `str(...)` boundary lives inside `parse_dates`, so a numeric member can
    legitimately fall back through date-syntax parsing. Malformed / out-of-range /
    reversed inputs surface as {resolved: false} or an empty resolved set."""
    try:
        # `DateRange` construction is inside the failure envelope: an unsupported
        # calendar year (e.g. ISO `0000`, which has no `datetime.date`) raises here
        # and must surface as a data-level unresolved result, not crash the oracle
        # process (qq0.23a fixup, closure-review P1).
        date_range = DateRange(startDate=req["startDate"], endDate=req["endDate"])
        map_did_d = _build_date_index_map(date_range, req.get("groups", []))
    except Exception as e:  # noqa: BLE001
        return {"ok": True, "resolved": False, "error": str(e)}
    try:
        indices = parse_dates(req["selector"], map_did_d, date_range)
        return {"ok": True, "resolved": True, "indices": indices}
    except Exception as e:  # noqa: BLE001
        return {"ok": True, "resolved": False, "error": str(e)}


def op_export(req):
    """C5: drive the real exporter. Accept -> CSV grid + xlsx byte length; a bad
    export config raises during export and surfaces as {ok: false}."""
    try:
        df, _sol, _score, status, cell = nurse_scheduling.schedule(
            req["yaml"].encode("utf-8"), prettify=True
        )
        if df is None:
            return {"ok": True, "status": status, "csv": None}
        buffer = io.BytesIO()
        exporter.export_to_excel(df, buffer, cell)
        # With prettify the frame is a pandas Styler; its underlying grid is `.data`.
        grid = df.data if hasattr(df, "data") else df
        # Re-open the produced workbook and extract exact cells/styles/notes so the
        # C5 harness can assert workbook contents, not just a byte count.
        buffer.seek(0)
        wb = load_workbook(buffer)
        ws = wb.active
        fills = {}
        cells = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    cells[c.coordinate] = c.value
                rgb = getattr(getattr(c.fill, "fgColor", None), "rgb", None)
                # openpyxl's unset fill is ARGB '00000000'; record anything else.
                if isinstance(rgb, str) and rgb not in ("00000000", None):
                    fills[c.coordinate] = rgb
        notes = None
        if "Notes" in wb.sheetnames:
            notes_ws = wb["Notes"]
            notes = [[cell.value for cell in nrow] for nrow in notes_ws.iter_rows()]
        return {
            "ok": True,
            "status": status,
            "csv": grid.to_csv(index=False, header=False, lineterminator="\n"),
            "xlsxBytes": len(buffer.getvalue()),
            "cells": cells,
            "fills": fills,
            "notes": notes,
            "frozen": ws.freeze_panes,
        }
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "error": str(e), "errorType": type(e).__name__}


def _atomize(model_dump):
    """Canonicalize a model dump for semantic round-trip comparison.

    Shift-request preferences are expanded into an order-independent multiset of
    (person, date, shiftType, weight) atoms — so the JS normalizer's neutral
    person/date/shiftType expansion into per-cell requests compares equal, while a
    dropped default or a flattened nested selector in ANY preference still diverges.
    Every other preference and all entities/dates/export are compared structurally.
    """
    prefs = model_dump.get("preferences", [])
    request_atoms = []
    other_prefs = []
    for pref in prefs:
        # Canonicalize implicit-all before structural compare: the producer emits
        # explicit `ALL` for an omitted requirement qualifiedPeople/date, which the
        # backend treats identically to `None`. Normalize both sides to `ALL` so a
        # legitimately-equivalent round-trip is not flagged as divergent.
        if pref.get("type") == "shift type requirement":
            pref = dict(pref)
            if pref.get("qualifiedPeople") is None:
                pref["qualifiedPeople"] = "ALL"
            if pref.get("date") is None:
                pref["date"] = "ALL"
        if pref.get("type") == "shift request":
            persons = pref["person"] if isinstance(pref["person"], list) else [pref["person"]]
            dates = pref["date"] if isinstance(pref["date"], list) else [pref["date"]]
            selectors = pref["shiftType"] if isinstance(pref["shiftType"], list) else [pref["shiftType"]]
            weight = pref.get("weight", 1)
            for p in persons:
                for d in dates:
                    for s in selectors:
                        request_atoms.append(json.dumps([p, str(d), s, weight], sort_keys=True, default=str))
        else:
            other_prefs.append(json.dumps(pref, sort_keys=True, default=str))
    canonical = dict(model_dump)
    canonical["preferences"] = {
        "requestAtoms": sorted(request_atoms),
        "otherPrefs": sorted(other_prefs),
    }
    return canonical


def op_roundtrip(req):
    """Import round-trip: load raw YAML and the JS-round-tripped YAML, then assert
    semantic-model-projection equivalence (raw -> Python vs
    raw -> import -> normalize -> canonical -> YAML -> Python).

    `appVersion` is frontend provenance metadata, not scenario semantics (C1
    CON-YAML-03; FR-SL-02 — every save re-stamps the current build version), so
    it is excluded from the comparison and returned separately for explicit
    assertion. Every other model field stays under strict comparison.
    """
    try:
        raw = _atomize(_json_safe(_dump(load_data(req["raw"].encode("utf-8")))))
        trip = _atomize(_json_safe(_dump(load_data(req["roundtrip"].encode("utf-8")))))
        app_version = {"raw": raw.pop("appVersion", None), "roundtrip": trip.pop("appVersion", None)}
        return {
            "ok": True,
            "equivalent": raw == trip,
            "raw": raw,
            "roundtrip": trip,
            "appVersion": app_version,
        }
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "error": str(e), "errorType": type(e).__name__}


def op_workspace_canonical(req):
    """T17r: drive the real pre-job submission boundary (`canonicalize_submission`).

    Accept -> the canonical strict YAML text plus the reparsed strict model dump;
    a scheduling-content rejection returns the normative envelope's error code and
    issues; malformed source returns the 400 marker. This is the authoritative
    Python half of the Workspace V1 cross-language contract.
    """
    try:
        canonical = canonicalize_submission(req["yaml"].encode("utf-8"))
        return {"ok": True, "canonical": canonical.decode("utf-8"), "model": _dump(load_data(canonical))}
    except MalformedInputError as e:
        return {"ok": False, "error": str(e), "errorType": "MalformedInputError"}
    except SchedulingContentError as e:
        return {
            "ok": False,
            "errorCode": e.error_code,
            "error": e.message,
            "issues": [issue.as_dict() for issue in e.issues],
        }


def op_workspace_equiv(req):
    """T17r: prove a TypeScript-produced Workspace document projects, through the
    real Python `canonicalize_submission`, to the same strict scheduling model as
    the frontend's own strict producer projection.

    `strict` is strict YAML loaded directly; `workspace` is Workspace V1 YAML run
    through the pre-job boundary. Both strict models are atomized (order-independent
    shift-request multiset) and compared. `appVersion` is provenance and is excluded
    from the comparison, returned separately for explicit assertion.
    """
    try:
        strict = _atomize(_json_safe(_dump(load_data(req["strict"].encode("utf-8")))))
        canonical = canonicalize_submission(req["workspace"].encode("utf-8"))
        workspace = _atomize(_json_safe(_dump(load_data(canonical))))
        app_version = {
            "strict": strict.pop("appVersion", None),
            "workspace": workspace.pop("appVersion", None),
        }
        return {
            "ok": True,
            "equivalent": strict == workspace,
            "strict": strict,
            "workspace": workspace,
            "appVersion": app_version,
        }
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "error": str(e), "errorType": type(e).__name__}


OPS = {
    "load": op_load,
    "schedule": op_schedule,
    "shift_map": op_shift_map,
    "people_map": op_people_map,
    "resolve_people": op_resolve_people,
    "resolve_dates": op_resolve_dates,
    "export": op_export,
    "roundtrip": op_roundtrip,
    "workspace_canonical": op_workspace_canonical,
    "workspace_equiv": op_workspace_equiv,
}


def main():
    request = json.load(sys.stdin)
    op = request.get("op")
    handler = OPS.get(op)
    if handler is None:
        print(json.dumps({"ok": False, "error": f"Unknown op: {op!r}"}))
        sys.exit(2)
    print(json.dumps(_json_safe(handler(request)), allow_nan=False))


if __name__ == "__main__":
    main()
