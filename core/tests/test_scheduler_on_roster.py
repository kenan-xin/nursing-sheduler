"""Tests for the additive `on_roster` day-state callback on `scheduler.schedule`.

The callback is the authoritative structured roster handoff: ordered typed
people, expanded ISO dates, an immutable single-state `solvedDays` grid, and
explicit 1-based worksheet axes. These tests pin the contract that downstream
consumers rely on, including a real-scenario coordinate-parity gate that fails
visibly if the exporter's layout ever drifts from the emitted coordinates.
"""

# This file is part of Nurse Scheduling Project, see <https://github.com/j3soon/nurse-scheduling>.
#
# Copyright (C) 2023-2026 Johnson Sun
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.

import json
import os
import sys
from io import BytesIO

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook

import nurse_scheduling
from nurse_scheduling import exporter
from nurse_scheduling.solver_interface import SolverStatus

CURRENT_DIR = os.path.dirname(os.path.realpath(__file__))
TESTCASES_DIR = f"{CURRENT_DIR}/testcases"

# A real 14-nurse / 28-day Singapore compliance scenario with five history
# entries per nurse, so prettify puts real history columns between the name
# column and the first date column.
REAL_SCENARIO = f"{TESTCASES_DIR}/real/sg-28day-160h-compliance-14-nurses.yaml"
HISTORY_SCENARIO = f"{TESTCASES_DIR}/basics/02_4nurses_3shifts_3days_unwanted_patterns_off_off_with_history.yaml"
INFEASIBLE_SCENARIO = f"{TESTCASES_DIR}/basics/01_1nurse_1shift_1day_infeasible.yaml"


def _read(path: str) -> bytes:
    with open(path, "rb") as f:
        return f.read()


def _capture(file_content: bytes, **kwargs):
    """Run a schedule, returning its five-value result plus captured payloads."""
    payloads = []
    result = nurse_scheduling.schedule(file_content, on_roster=payloads.append, **kwargs)
    return result, payloads


def _display(day: dict) -> str:
    """The exporter's cell rendering for a day-state."""
    if day["kind"] == "off":
        return ""
    if day["kind"] == "leave":
        return "Leave"
    return day["shiftId"]


# A worked / OFF / Leave scenario: nurse 0 is hard-pinned to paid leave on the
# second date, exactly one nurse works D every date, so every day-state kind
# appears in the grid.
DAY_STATE_SCENARIO = b"""
apiVersion: alpha
description: One worked shift per date, one pinned leave, the rest OFF
dates:
  range:
    startDate: 2026-02-01
    endDate: 2026-02-04
people:
  items:
    - id: 0
    - id: 1
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    description: Exactly one nurse on D every date
    shiftType: D
    requiredNumPeople: 1
  - type: shift request
    description: Nurse 0 on paid leave Feb 2 (hard pin)
    person: 0
    date: 2
    shiftType: LEAVE
    weight: .inf
"""


def test_callback_reports_worked_off_and_leave_day_states():
    (df, _solution, _score, status, _cell_export_info), payloads = _capture(DAY_STATE_SCENARIO)

    assert status in ("OPTIMAL", "FEASIBLE")
    assert len(payloads) == 1
    payload = payloads[0]

    solved_days = payload["solvedDays"]
    assert len(solved_days) == 2
    assert all(len(person_days) == 4 for person_days in solved_days)

    kinds = {day["kind"] for person_days in solved_days for day in person_days}
    assert kinds == {"shift", "off", "leave"}

    # The pinned leave lands on exactly the requested cell, and nowhere else.
    assert solved_days[0][1] == {"kind": "leave"}
    leave_cells = [
        (p, d)
        for p, person_days in enumerate(solved_days)
        for d, day in enumerate(person_days)
        if day["kind"] == "leave"
    ]
    assert leave_cells == [(0, 1)]

    # A worked state is a single shift id, never a list.
    for person_days in solved_days:
        for day in person_days:
            if day["kind"] == "shift":
                assert day["shiftId"] == "D"
                assert set(day) == {"kind", "shiftId"}
            else:
                assert set(day) == {"kind"}

    # Exactly one nurse works D on each date; the other is OFF or on leave.
    for d in range(4):
        worked = [p for p in range(2) if solved_days[p][d]["kind"] == "shift"]
        assert worked == [1] if d == 1 else len(worked) == 1

    # Day-states agree with the dataframe the same run exported.
    for p, person_days in enumerate(solved_days):
        for d, day in enumerate(person_days):
            assert df.iloc[2 + p, 1 + d] == _display(day)


def test_callback_fires_for_a_feasible_but_not_optimal_outcome(monkeypatch):
    from nurse_scheduling.solver_ortools_cp_sat import ORToolsSolver

    real_solve = ORToolsSolver.solve

    def solve_as_feasible(self, *args, **kwargs):
        real_solve(self, *args, **kwargs)
        return SolverStatus.FEASIBLE

    # The solved values stay real; only the reported outcome is downgraded, so
    # this exercises the FEASIBLE branch deterministically.
    monkeypatch.setattr(ORToolsSolver, "solve", solve_as_feasible)
    monkeypatch.setattr(ORToolsSolver, "get_status_name", lambda self: "FEASIBLE")

    (df, _solution, _score, status, _cell_export_info), payloads = _capture(DAY_STATE_SCENARIO)

    assert status == "FEASIBLE"
    assert df is not None
    assert len(payloads) == 1
    assert {day["kind"] for person_days in payloads[0]["solvedDays"] for day in person_days} == {
        "shift",
        "off",
        "leave",
    }


def test_callback_is_skipped_when_no_solution_exists():
    (df, solution, score, status, cell_export_info), payloads = _capture(_read(INFEASIBLE_SCENARIO))

    assert status == "INFEASIBLE"
    assert (df, solution, score, cell_export_info) == (None, None, None, None)
    assert payloads == []


def test_return_shape_is_unchanged_with_and_without_the_callback():
    without = nurse_scheduling.schedule(DAY_STATE_SCENARIO)
    with_callback, payloads = _capture(DAY_STATE_SCENARIO)

    assert len(without) == 5
    assert len(with_callback) == 5
    assert len(payloads) == 1

    df_a, solution_a, score_a, status_a, cell_export_info_a = without
    df_b, solution_b, score_b, status_b, cell_export_info_b = with_callback
    assert df_a.equals(df_b)
    assert solution_a == solution_b
    assert (score_a, status_a, cell_export_info_a) == (score_b, status_b, cell_export_info_b)


def test_on_roster_is_keyword_only():
    import inspect

    parameter = inspect.signature(nurse_scheduling.schedule).parameters["on_roster"]
    assert parameter.kind is inspect.Parameter.KEYWORD_ONLY
    assert parameter.default is None


def test_axes_are_ordered_and_preserve_the_authored_id_type():
    _result, payloads = _capture(DAY_STATE_SCENARIO)
    payload = payloads[0]

    # Integer ids stay integers, and survive a JSON container roundtrip as such.
    assert payload["people"] == [{"id": 0}, {"id": 1}]
    assert json.loads(json.dumps(payload))["people"] == [{"id": 0}, {"id": 1}]

    assert payload["dates"] == [
        {"iso": "2026-02-01"},
        {"iso": "2026-02-02"},
        {"iso": "2026-02-03"},
        {"iso": "2026-02-04"},
    ]
    assert len(payload["solvedDays"]) == len(payload["people"])
    assert all(len(row) == len(payload["dates"]) for row in payload["solvedDays"])

    _result, real_payloads = _capture(_read(REAL_SCENARIO), prettify=True)
    real_payload = real_payloads[0]
    # String ids stay strings, in the authored order.
    assert [person["id"] for person in real_payload["people"]] == [f"N{n:02d}" for n in range(1, 15)]


def test_coordinate_map_tracks_history_columns_and_prettify():
    file_content = _read(HISTORY_SCENARIO)

    # Without prettify the exporter emits no history columns, so dates start
    # immediately after the single name column.
    _result, plain = _capture(file_content)
    plain_map = plain[0]["coordinateMap"]
    assert plain_map["prettify"] is False
    assert plain_map["historyCols"] == 0
    assert plain_map["leadingCols"] == 1
    assert plain_map["firstPeopleRow"] == 3
    assert plain_map["peopleRows"] == [3, 4, 5, 6]
    assert plain_map["dateColumns"] == [2, 3, 4]

    # With prettify the globally longest history (1 entry here) inserts one
    # history column, shifting every date column right by one.
    _result, pretty = _capture(file_content, prettify=True)
    pretty_map = pretty[0]["coordinateMap"]
    assert pretty_map["prettify"] is True
    assert pretty_map["historyCols"] == 1
    assert pretty_map["peopleRows"] == [3, 4, 5, 6]
    assert pretty_map["dateColumns"] == [3, 4, 5]

    # The real scenario's five-entry histories give five history columns.
    _result, real = _capture(_read(REAL_SCENARIO), prettify=True)
    real_map = real[0]["coordinateMap"]
    assert real_map["historyCols"] == 5
    assert real_map["dateColumns"][0] == 7


def _assert_callback_matches_workbook(file_content: bytes, prettify: bool) -> None:
    """Coordinate-parity gate: every roster cell of the callback handoff must
    land on the matching cell of the workbook the same run exported.

    The payload is passed through a container-shaped JSON roundtrip first, so
    the gate proves what a downstream JSON container consumer would actually
    receive. B2 owns the production container; this stays test-local.
    """
    payloads = []
    df, _solution, _score, status, cell_export_info = nurse_scheduling.schedule(
        file_content,
        prettify=prettify,
        on_roster=payloads.append,
    )
    assert status in ("OPTIMAL", "FEASIBLE")
    assert len(payloads) == 1

    container = json.loads(
        json.dumps(
            {
                "schemaVersion": "roster-container/1",
                "people": payloads[0]["people"],
                "dates": payloads[0]["dates"],
                "solvedDays": payloads[0]["solvedDays"],
                "coordinateMap": payloads[0]["coordinateMap"],
            }
        )
    )

    output = BytesIO()
    exporter.export_to_excel(df, output, cell_export_info)
    worksheet = load_workbook(BytesIO(output.getvalue())).worksheets[0]

    coordinate_map = container["coordinateMap"]
    people_rows = coordinate_map["peopleRows"]
    date_columns = coordinate_map["dateColumns"]
    assert len(people_rows) == len(container["people"])
    assert len(date_columns) == len(container["dates"])

    for person_idx, person_days in enumerate(container["solvedDays"]):
        for date_idx, day in enumerate(person_days):
            cell = worksheet.cell(row=people_rows[person_idx], column=date_columns[date_idx])
            actual = "" if cell.value is None else cell.value
            assert actual == _display(day), (
                f"Coordinate parity drift at person {container['people'][person_idx]['id']} / "
                f"{container['dates'][date_idx]['iso']} "
                f"(row {people_rows[person_idx]}, column {date_columns[date_idx]})"
            )

    # The people window really does start where the map says, and the name
    # column still holds the person ids in payload order.
    assert people_rows[0] == coordinate_map["firstPeopleRow"]
    for person_idx, person in enumerate(container["people"]):
        assert str(worksheet.cell(row=people_rows[person_idx], column=1).value) == str(person["id"])

    # Date columns sit immediately after the name + history columns.
    assert date_columns[0] == coordinate_map["leadingCols"] + coordinate_map["historyCols"] + 1
    assert date_columns == list(range(date_columns[0], date_columns[0] + len(date_columns)))


def test_callback_to_container_coordinate_parity_on_the_real_scenario_with_history():
    _assert_callback_matches_workbook(_read(REAL_SCENARIO), prettify=True)


def test_callback_to_container_coordinate_parity_without_prettify():
    _assert_callback_matches_workbook(_read(REAL_SCENARIO), prettify=False)


def test_callback_to_container_coordinate_parity_covers_leave_cells():
    _assert_callback_matches_workbook(DAY_STATE_SCENARIO, prettify=True)
    _assert_callback_matches_workbook(DAY_STATE_SCENARIO, prettify=False)
