"""Tests for the `roster-container/1` artifact built by `OptimizationRunner`.

The container is the single stored artifact of a completed job, so these tests
pin the two properties everything downstream rests on: the embedded workbook is
byte-identical to what the exporter produced, and the structured axes address
exactly the cells of that workbook. The frozen size caps are proved at their
boundaries, including that an oversized run never puts bytes on the child result
pipe.
"""

# This file is part of Nurse Scheduling Project, see <https://github.com/j3soon/nurse-scheduling>.
#
# Copyright (C) 2023-2026 Johnson Sun
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.

import base64
import hashlib
import json
import os
from datetime import UTC, datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook

from nurse_scheduling import exporter
from nurse_scheduling.server.errors import OptimizationExecutionError
from nurse_scheduling.server.jobs.models import Job, JobRequest, JobState
from nurse_scheduling.server.jobs.process_executor import ProcessStatus, run_optimization_process
from nurse_scheduling.server.jobs.runner import OptimizationRunner
from nurse_scheduling.server.roster_container import (
    INVALID_OUTPUT_CODE,
    MAX_RAW_XLSX_BYTES,
    MAX_ROSTER_CONTAINER_BYTES,
    SCHEMA_VERSION,
    XLSX_MEDIA_TYPE,
    RosterContainerInvalidError,
    build_roster_container,
    decode_workbook,
    parse_roster_container,
    roster_view,
    workbook_download_name,
    workbook_media_type,
)

CURRENT_DIR = os.path.dirname(os.path.realpath(__file__))
TESTCASES_DIR = f"{CURRENT_DIR}/testcases"

# The same real 14-nurse / 28-day Singapore compliance scenario B1's parity gate
# uses: five history entries per nurse, so prettify puts real history columns
# between the name column and the first date column.
REAL_SCENARIO = f"{TESTCASES_DIR}/real/sg-28day-160h-compliance-14-nurses.yaml"

# Reused by the boundary tests: a structurally valid handoff whose encoded size
# is dominated by the embedded workbook.
ROSTER_PAYLOAD = {
    "people": [{"id": 0}, {"id": "N02"}],
    "dates": [{"iso": "2026-02-01"}, {"iso": "2026-02-02"}],
    "solvedDays": [
        [{"kind": "shift", "shiftId": "D"}, {"kind": "off"}],
        [{"kind": "leave"}, {"kind": "shift", "shiftId": "D"}],
    ],
    "coordinateMap": {
        "peopleRows": [3, 4],
        "dateColumns": [2, 3],
        "firstPeopleRow": 3,
        "leadingCols": 1,
        "historyCols": 0,
        "prettify": False,
    },
}


def _build(xlsx_bytes: bytes, **limits) -> bytes:
    """Build a container over fixed structure, varying only bytes and limits."""
    return build_roster_container(
        ROSTER_PAYLOAD,
        xlsx_bytes=xlsx_bytes,
        xlsx_name="nurse-scheduling-20260201T000000Z.xlsx",
        xlsx_mime=XLSX_MEDIA_TYPE,
        score=42,
        solver_status="OPTIMAL",
        **limits,
    )


def _encode(container: dict) -> bytes:
    """Serialize a (possibly mutated) container the way the store holds it."""
    return json.dumps(container, separators=(",", ":")).encode("utf-8")


def _valid_container() -> dict:
    """A freshly parsed container that satisfies the whole v1 contract."""
    return json.loads(_build(b"workbook"))


def _valid_handoff() -> dict:
    """The four scheduler-owned fields, detached from a valid container."""
    container = _valid_container()
    return {field: container[field] for field in ("people", "dates", "solvedDays", "coordinateMap")}


def _job(prettify: bool) -> Job:
    return Job(
        id="job_roster_container",
        state=JobState.RUNNING,
        request=JobRequest(
            input_name="scenario.yaml",
            client_id="client",
            solver="ortools/cp-sat",
            prettify=prettify,
            timeout_seconds=300,
        ),
        created_at=datetime.now(UTC),
    )


def run_with_export_spy(monkeypatch, scenario_path: str, *, prettify: bool) -> tuple[dict, bytes, bytes]:
    """Run the production runner, returning the container, its bytes, and the exporter's bytes.

    The spy wraps the real exporter, so the recorded bytes are exactly what was
    written to the workbook buffer before the container encoded them.
    """
    exported: list[bytes] = []
    real_export = exporter.export_to_excel

    def spying_export(dataframe, buffer, cell_export_info):
        real_export(dataframe, buffer, cell_export_info)
        exported.append(buffer.getvalue())

    monkeypatch.setattr("nurse_scheduling.exporter.export_to_excel", spying_export)
    with open(scenario_path, "rb") as scenario:
        input_bytes = scenario.read()

    output = OptimizationRunner().run(
        _job(prettify),
        input_bytes,
        event_callback=lambda *_args: None,
        should_stop=None,
    )

    assert output.artifact is not None
    assert len(exported) == 1
    return parse_roster_container(output.artifact.content), output.artifact.content, exported[0]


def _display(day: dict) -> str:
    """The exporter's cell rendering for a day-state."""
    if day["kind"] == "off":
        return ""
    if day["kind"] == "leave":
        return "Leave"
    return day["shiftId"]


# --- Frozen production constants ---------------------------------------------


def test_frozen_size_caps_are_the_agreed_v1_byte_values():
    assert MAX_RAW_XLSX_BYTES == 33_554_432
    assert MAX_ROSTER_CONTAINER_BYTES == 50_331_648


# --- Base64 / hash parity -----------------------------------------------------


def test_embedded_workbook_is_hash_identical_to_the_exporter_bytes(monkeypatch):
    container, _stored, exported = run_with_export_spy(monkeypatch, REAL_SCENARIO, prettify=True)

    decoded = decode_workbook(container)
    assert decoded == exported
    assert hashlib.sha256(decoded).hexdigest() == hashlib.sha256(exported).hexdigest()
    # Strict base64: the stored payload has no whitespace or alien characters.
    assert base64.b64decode(container["xlsx"]["base64"], validate=True) == exported
    assert container["xlsx"]["mime"] == XLSX_MEDIA_TYPE
    assert container["xlsx"]["name"].endswith(".xlsx")


# --- End-to-end callback -> container -> real workbook ------------------------


@pytest.mark.parametrize("prettify", [True, False])
def test_real_scenario_container_addresses_every_workbook_coordinate(monkeypatch, prettify):
    container, _stored, _exported = run_with_export_spy(monkeypatch, REAL_SCENARIO, prettify=prettify)
    worksheet = load_workbook(BytesIO(decode_workbook(container))).worksheets[0]

    assert container["schemaVersion"] == SCHEMA_VERSION
    assert container["solverStatus"] in ("OPTIMAL", "FEASIBLE")
    assert isinstance(container["score"], (int, float))

    # Ordered typed axes survived the container roundtrip: authored string ids in
    # authored order, and expanded ISO dates in calendar order.
    assert [person["id"] for person in container["people"]] == [f"N{n:02d}" for n in range(1, 15)]
    assert [date["iso"] for date in container["dates"]] == sorted(date["iso"] for date in container["dates"])
    assert len(container["dates"]) == 28

    coordinate_map = container["coordinateMap"]
    people_rows = coordinate_map["peopleRows"]
    date_columns = coordinate_map["dateColumns"]
    assert coordinate_map["prettify"] is prettify
    assert coordinate_map["historyCols"] == (5 if prettify else 0)
    assert len(people_rows) == len(container["people"])
    assert len(date_columns) == len(container["dates"])
    assert date_columns[0] == coordinate_map["leadingCols"] + coordinate_map["historyCols"] + 1

    # The grid is complete and single-state, and every emitted coordinate lands
    # on the matching cell of the workbook the same run exported.
    assert len(container["solvedDays"]) == len(container["people"])
    for person_idx, person_days in enumerate(container["solvedDays"]):
        assert len(person_days) == len(container["dates"])
        assert str(worksheet.cell(row=people_rows[person_idx], column=1).value) == str(
            container["people"][person_idx]["id"]
        )
        for date_idx, day in enumerate(person_days):
            if day["kind"] == "shift":
                assert set(day) == {"kind", "shiftId"}
            else:
                assert set(day) == {"kind"}
            cell = worksheet.cell(row=people_rows[person_idx], column=date_columns[date_idx])
            actual = "" if cell.value is None else cell.value
            assert actual == _display(day), (
                f"Coordinate drift at {container['people'][person_idx]['id']} / "
                f"{container['dates'][date_idx]['iso']} "
                f"(row {people_rows[person_idx]}, column {date_columns[date_idx]})"
            )


# --- Size caps at their boundaries -------------------------------------------


def test_raw_workbook_cap_accepts_the_limit_and_rejects_one_byte_more():
    limit = 4096

    accepted = _build(b"x" * limit, max_raw_xlsx_bytes=limit, max_container_bytes=MAX_ROSTER_CONTAINER_BYTES)
    assert base64.b64decode(json.loads(accepted)["xlsx"]["base64"], validate=True) == b"x" * limit

    with pytest.raises(OptimizationExecutionError) as raised:
        _build(b"x" * (limit + 1), max_raw_xlsx_bytes=limit, max_container_bytes=MAX_ROSTER_CONTAINER_BYTES)
    assert raised.value.code == "roster_output_too_large"
    assert f"{limit + 1} bytes" in str(raised.value)


def test_encoded_container_cap_accepts_the_limit_and_rejects_one_byte_more():
    # Measure the real encoded size first, then treat it as the limit: the same
    # document is accepted at exactly the limit and rejected when the limit is
    # one byte lower, which is the limit + 1 case from the cap's point of view.
    workbook = b"x" * 4096
    encoded_size = len(_build(workbook))

    accepted = _build(workbook, max_container_bytes=encoded_size)
    assert len(accepted) == encoded_size

    with pytest.raises(OptimizationExecutionError) as raised:
        _build(workbook, max_container_bytes=encoded_size - 1)
    assert raised.value.code == "roster_output_too_large"
    assert f"container is {encoded_size} bytes" in str(raised.value)


class OversizedRosterRunner:
    """Fails the frozen encoded cap while producing a valid roster handoff.

    Defined at module level so the spawn-based executor can pickle it.
    """

    def run(self, job, input_bytes, *, event_callback, should_stop):
        build_roster_container(
            ROSTER_PAYLOAD,
            xlsx_bytes=b"x" * 4096,
            xlsx_name="nurse-scheduling-20260201T000000Z.xlsx",
            xlsx_mime=XLSX_MEDIA_TYPE,
            score=42,
            solver_status="OPTIMAL",
            max_container_bytes=512,
        )
        raise AssertionError("the oversized container must not be built")


class InvalidHandoffRunner:
    """Mimics a regressed scheduler handoff: a grid row that is not rectangular.

    Defined at module level so the spawn-based executor can pickle it.
    """

    def run(self, job, input_bytes, *, event_callback, should_stop):
        payload = _valid_handoff()
        payload["solvedDays"][0].pop()
        build_roster_container(
            payload,
            xlsx_bytes=b"workbook",
            xlsx_name="nurse-scheduling-20260201T000000Z.xlsx",
            xlsx_mime=XLSX_MEDIA_TYPE,
            score=42,
            solver_status="OPTIMAL",
        )
        raise AssertionError("the invalid container must not be built")


class GappedAxisRunner:
    """Mimics a regressed scheduler handoff: a non-first row-axis gap.

    Defined at module level so the spawn-based executor can pickle it.
    """

    def run(self, job, input_bytes, *, event_callback, should_stop):
        payload = _valid_handoff()
        payload["coordinateMap"]["peopleRows"] = [
            payload["coordinateMap"]["firstPeopleRow"],
            payload["coordinateMap"]["firstPeopleRow"] + 2,
        ]
        build_roster_container(
            payload,
            xlsx_bytes=b"workbook",
            xlsx_name="nurse-scheduling-20260201T000000Z.xlsx",
            xlsx_mime=XLSX_MEDIA_TYPE,
            score=42,
            solver_status="OPTIMAL",
        )
        raise AssertionError("the gapped container must not be built")


class GappedDateColumnsAxisRunner:
    """Mimics a regressed scheduler handoff: a non-first date-column-axis gap.

    Defined at module level so the spawn-based executor can pickle it.
    """

    def run(self, job, input_bytes, *, event_callback, should_stop):
        payload = _valid_handoff()
        coordinate_map = payload["coordinateMap"]
        first_column = coordinate_map["leadingCols"] + coordinate_map["historyCols"] + 1
        coordinate_map["dateColumns"] = [first_column, first_column + 2]
        build_roster_container(
            payload,
            xlsx_bytes=b"workbook",
            xlsx_name="nurse-scheduling-20260201T000000Z.xlsx",
            xlsx_mime=XLSX_MEDIA_TYPE,
            score=42,
            solver_status="OPTIMAL",
        )
        raise AssertionError("the gapped container must not be built")


def test_an_invalid_handoff_never_crosses_the_child_result_pipe():
    # A contract violation terminates the child with the stable code and no
    # output, so an unusable container cannot be pickled to the supervisor.
    result = run_optimization_process(
        InvalidHandoffRunner(),
        _job(prettify=False),
        b"apiVersion: alpha\n",
        event_callback=lambda *_args: None,
        control=lambda: None,
        hard_timeout_seconds=61,
        finish_now_enabled=False,
    )

    assert result.status is ProcessStatus.FAILED
    assert result.output is None
    assert result.failure is not None
    assert result.failure.code == INVALID_OUTPUT_CODE


def test_a_non_first_axis_gap_never_crosses_the_child_result_pipe():
    # A strictly-increasing-but-gapped axis is just as much an internal
    # regression as a ragged grid: it must fail before any output exists rather
    # than pickling an unusable container to the supervisor.
    result = run_optimization_process(
        GappedAxisRunner(),
        _job(prettify=False),
        b"apiVersion: alpha\n",
        event_callback=lambda *_args: None,
        control=lambda: None,
        hard_timeout_seconds=61,
        finish_now_enabled=False,
    )

    assert result.status is ProcessStatus.FAILED
    assert result.output is None
    assert result.failure is not None
    assert result.failure.code == INVALID_OUTPUT_CODE


def test_oversized_output_never_crosses_the_child_result_pipe():
    # The child sends only the stable terminal failure: no RunOutput exists, so
    # no container or workbook bytes are pickled onto the result pipe.
    result = run_optimization_process(
        OversizedRosterRunner(),
        _job(prettify=False),
        b"apiVersion: alpha\n",
        event_callback=lambda *_args: None,
        control=lambda: None,
        hard_timeout_seconds=61,
        finish_now_enabled=False,
    )

    assert result.status is ProcessStatus.FAILED
    assert result.output is None
    assert result.failure is not None
    assert result.failure.code == "roster_output_too_large"


# --- Parsing and projection ---------------------------------------------------


def test_roster_view_removes_only_the_embedded_workbook_bytes():
    container = json.loads(_build(b"workbook"))

    view = roster_view(container)

    assert "base64" not in view["xlsx"]
    assert view["xlsx"] == {"name": container["xlsx"]["name"], "mime": container["xlsx"]["mime"]}
    assert {key: value for key, value in view.items() if key != "xlsx"} == {
        key: value for key, value in container.items() if key != "xlsx"
    }
    # The projection does not mutate the parsed container it was given.
    assert "base64" in container["xlsx"]


def test_container_encoding_is_deterministic_and_utf8_json():
    first = _build(b"workbook")
    second = _build(b"workbook")

    assert first == second
    assert first.decode("utf-8").startswith('{"')
    # Deterministic means sorted keys and no insignificant whitespace.
    assert b", " not in first
    assert first.decode("utf-8").index('"coordinateMap"') < first.decode("utf-8").index('"dates"')


def test_a_well_formed_container_round_trips_through_the_validator():
    container = parse_roster_container(_build(b"workbook"))

    assert container["schemaVersion"] == SCHEMA_VERSION
    assert decode_workbook(container) == b"workbook"


@pytest.mark.parametrize(
    "content",
    [
        pytest.param(b"not json at all", id="not-json"),
        pytest.param(b"[]", id="json-array"),
        pytest.param(b'"a string"', id="json-string"),
        pytest.param(b"\xff\xfe not utf-8", id="not-utf8"),
        pytest.param(b'{"schemaVersion":"roster-container/1"}', id="only-schema-version"),
        pytest.param(b'{"schemaVersion":"roster-container/1","xlsx":{"base64":""}}', id="schema-version-and-workbook"),
    ],
)
def test_unreadable_container_bytes_fail_closed(content):
    with pytest.raises(RosterContainerInvalidError):
        parse_roster_container(content)


@pytest.mark.parametrize("literal", ["NaN", "Infinity", "-Infinity"])
def test_non_standard_json_constants_are_rejected(literal):
    # Python's `json` accepts these by default; a `roster-container/1` document is
    # strict standard JSON, so a score of NaN must never reach a consumer.
    container = _valid_container()
    content = _encode(container).replace(b'"score":42', f'"score":{literal}'.encode())

    with pytest.raises(RosterContainerInvalidError, match="non-standard JSON constant"):
        parse_roster_container(content)


def test_an_overflowing_json_float_is_rejected_as_non_finite():
    # `1e400` is standard JSON syntax that Python decodes to infinity.
    content = _encode(_valid_container()).replace(b'"score":42', b'"score":1e400')

    with pytest.raises(RosterContainerInvalidError, match="finite number"):
        parse_roster_container(content)


# Field-by-field and cross-field rejections. Each mutation starts from a document
# that parses cleanly, so every case pins exactly one contract rule.
#
# Structural cases mutate the four fields the scheduler handoff owns, so they
# apply to the build path as well as the read path.
STRUCTURAL_VIOLATIONS = [
    ("people-not-an-array", lambda c: c.update({"people": {"0": {"id": 0}}})),
    ("person-record-with-extra-field", lambda c: c["people"][0].update({"description": "Alice"})),
    ("person-record-wrong-key", lambda c: c["people"].__setitem__(0, {"personId": 0})),
    ("person-id-boolean", lambda c: c["people"].__setitem__(0, {"id": True})),
    ("person-id-empty-string", lambda c: c["people"].__setitem__(0, {"id": ""})),
    ("person-id-object", lambda c: c["people"].__setitem__(0, {"id": {"value": 1}})),
    ("person-id-null", lambda c: c["people"].__setitem__(0, {"id": None})),
    ("dates-not-an-array", lambda c: c.update({"dates": "2026-02-01"})),
    ("date-record-wrong-key", lambda c: c["dates"].__setitem__(0, {"date": "2026-02-01"})),
    ("date-not-iso-shaped", lambda c: c["dates"].__setitem__(0, {"iso": "01/02/2026"})),
    ("date-not-a-calendar-date", lambda c: c["dates"].__setitem__(0, {"iso": "2026-02-30"})),
    ("date-with-time", lambda c: c["dates"].__setitem__(0, {"iso": "2026-02-01T08:00"})),
    ("grid-missing-a-person-row", lambda c: c["solvedDays"].pop()),
    ("grid-row-not-an-array", lambda c: c["solvedDays"].__setitem__(0, {"0": {"kind": "off"}})),
    ("ragged-grid-row", lambda c: c["solvedDays"][0].pop()),
    ("grid-cell-not-an-object", lambda c: c["solvedDays"][0].__setitem__(0, "D")),
    ("unknown-day-kind", lambda c: c["solvedDays"][0].__setitem__(0, {"kind": "holiday"})),
    ("worked-day-without-shift", lambda c: c["solvedDays"][0].__setitem__(0, {"kind": "shift"})),
    ("worked-day-with-shift-list", lambda c: c["solvedDays"][0].__setitem__(0, {"kind": "shift", "shiftId": ["D"]})),
    ("worked-day-with-empty-shift", lambda c: c["solvedDays"][0].__setitem__(0, {"kind": "shift", "shiftId": ""})),
    ("off-day-with-a-payload", lambda c: c["solvedDays"][0].__setitem__(1, {"kind": "off", "shiftId": "D"})),
    ("coordinate-map-missing-field", lambda c: c["coordinateMap"].pop("prettify")),
    ("coordinate-map-extra-field", lambda c: c["coordinateMap"].update({"trailingCols": 0})),
    ("people-rows-cardinality", lambda c: c["coordinateMap"].update({"peopleRows": [3]})),
    ("date-columns-cardinality", lambda c: c["coordinateMap"].update({"dateColumns": [2, 3, 4]})),
    ("people-row-not-positive", lambda c: c["coordinateMap"].update({"peopleRows": [3, -4]})),
    ("people-row-not-an-integer", lambda c: c["coordinateMap"].update({"peopleRows": [3, 4.5]})),
    ("people-rows-not-increasing", lambda c: c["coordinateMap"].update({"peopleRows": [4, 3]})),
    ("date-columns-not-increasing", lambda c: c["coordinateMap"].update({"dateColumns": [3, 2]})),
    ("people-rows-off-anchor", lambda c: c["coordinateMap"].update({"peopleRows": [5, 6]})),
    ("date-columns-off-anchor", lambda c: c["coordinateMap"].update({"dateColumns": [4, 5]})),
    # Strictly increasing but not contiguous: the real exporter can never skip a
    # row/column, so a non-first gap must be rejected even though the anchor and
    # ordering checks alone would let it through.
    ("people-rows-non-first-gap", lambda c: c["coordinateMap"].update({"peopleRows": [3, 5]})),
    ("date-columns-non-first-gap", lambda c: c["coordinateMap"].update({"dateColumns": [2, 4]})),
    ("first-people-row-zero", lambda c: c["coordinateMap"].update({"firstPeopleRow": 0})),
    ("first-people-row-not-an-integer", lambda c: c["coordinateMap"].update({"firstPeopleRow": "3"})),
    ("leading-cols-float", lambda c: c["coordinateMap"].update({"leadingCols": 1.0})),
    ("leading-cols-negative", lambda c: c["coordinateMap"].update({"leadingCols": -1})),
    ("history-cols-not-an-integer", lambda c: c["coordinateMap"].update({"historyCols": None})),
    ("history-cols-without-prettify", lambda c: c["coordinateMap"].update({"historyCols": 1})),
    ("prettify-not-a-boolean", lambda c: c["coordinateMap"].update({"prettify": "yes"})),
]

# Metadata cases mutate fields the builder itself supplies, so they are read-path
# rejections; the build path covers the same ground through its own arguments.
METADATA_VIOLATIONS = [
    ("missing-required-field", lambda c: c.pop("score")),
    ("missing-coordinate-map", lambda c: c.pop("coordinateMap")),
    ("unexpected-top-level-field", lambda c: c.update({"edits": []})),
    ("wrong-schema-version", lambda c: c.update({"schemaVersion": "roster-container/2"})),
    ("score-null", lambda c: c.update({"score": None})),
    ("score-string", lambda c: c.update({"score": "42"})),
    ("score-boolean", lambda c: c.update({"score": True})),
    ("status-not-a-solved-run", lambda c: c.update({"solverStatus": "INFEASIBLE"})),
    ("status-not-a-string", lambda c: c.update({"solverStatus": None})),
    ("workbook-not-an-object", lambda c: c.update({"xlsx": "base64"})),
    ("workbook-missing-mime", lambda c: c["xlsx"].pop("mime")),
    ("workbook-extra-field", lambda c: c["xlsx"].update({"bytes": 8})),
    ("workbook-wrong-mime", lambda c: c["xlsx"].update({"mime": "text/html"})),
    ("workbook-name-not-a-workbook", lambda c: c["xlsx"].update({"name": "schedule.csv"})),
    ("workbook-name-not-a-string", lambda c: c["xlsx"].update({"name": None})),
    ("workbook-base64-not-strict", lambda c: c["xlsx"].update({"base64": "not base64!"})),
    ("workbook-base64-not-a-string", lambda c: c["xlsx"].update({"base64": 123})),
]


@pytest.mark.parametrize(
    "mutate",
    [pytest.param(mutate, id=case_id) for case_id, mutate in STRUCTURAL_VIOLATIONS + METADATA_VIOLATIONS],
)
def test_contract_violations_are_rejected_on_read(mutate):
    container = _valid_container()
    mutate(container)

    with pytest.raises(RosterContainerInvalidError):
        parse_roster_container(_encode(container))


@pytest.mark.parametrize(
    "mutate",
    [pytest.param(mutate, id=case_id) for case_id, mutate in STRUCTURAL_VIOLATIONS],
)
def test_contract_violations_in_the_handoff_are_rejected_on_build(mutate):
    # The same shared validator guards the build path, so a regressed scheduler
    # handoff fails before any bytes exist.
    payload = _valid_handoff()
    mutate(payload)

    with pytest.raises(OptimizationExecutionError) as raised:
        build_roster_container(
            payload,
            xlsx_bytes=b"workbook",
            xlsx_name="nurse-scheduling-20260201T000000Z.xlsx",
            xlsx_mime=XLSX_MEDIA_TYPE,
            score=42,
            solver_status="OPTIMAL",
        )
    assert raised.value.code == INVALID_OUTPUT_CODE


@pytest.mark.parametrize(
    ("case_id", "overrides"),
    [
        ("non-finite-score", {"score": float("nan")}),
        ("infinite-score", {"score": float("inf")}),
        ("missing-score", {"score": None}),
        ("unsolved-status", {"solver_status": "INFEASIBLE"}),
        ("workbook-name-not-a-workbook", {"xlsx_name": "schedule.csv"}),
        ("workbook-mime-not-a-workbook", {"xlsx_mime": "text/html"}),
    ],
)
def test_invalid_build_metadata_fails_before_any_output(case_id, overrides):
    arguments = {
        "xlsx_bytes": b"workbook",
        "xlsx_name": "nurse-scheduling-20260201T000000Z.xlsx",
        "xlsx_mime": XLSX_MEDIA_TYPE,
        "score": 42,
        "solver_status": "OPTIMAL",
        **overrides,
    }

    with pytest.raises(OptimizationExecutionError) as raised:
        build_roster_container(ROSTER_PAYLOAD, **arguments)
    assert raised.value.code == INVALID_OUTPUT_CODE


@pytest.mark.parametrize(
    "payload",
    [
        pytest.param(None, id="handoff-is-none"),
        pytest.param([], id="handoff-is-a-list"),
        pytest.param({"people": [], "dates": []}, id="handoff-missing-grid-and-axes"),
    ],
)
def test_a_structurally_absent_handoff_fails_before_any_output(payload):
    with pytest.raises(OptimizationExecutionError) as raised:
        build_roster_container(
            payload,
            xlsx_bytes=b"workbook",
            xlsx_name="nurse-scheduling-20260201T000000Z.xlsx",
            xlsx_mime=XLSX_MEDIA_TYPE,
            score=42,
            solver_status="OPTIMAL",
        )
    assert raised.value.code == INVALID_OUTPUT_CODE


def test_the_raw_size_cap_is_still_checked_before_validation():
    # Ordering is part of the frozen behavior: an oversized workbook is reported
    # as too large even when the handoff is also malformed.
    with pytest.raises(OptimizationExecutionError) as raised:
        build_roster_container(
            {"people": [], "dates": [], "solvedDays": [[]], "coordinateMap": {}},
            xlsx_bytes=b"x" * 11,
            xlsx_name="nurse-scheduling-20260201T000000Z.xlsx",
            xlsx_mime=XLSX_MEDIA_TYPE,
            score=42,
            solver_status="OPTIMAL",
            max_raw_xlsx_bytes=10,
        )
    assert raised.value.code == "roster_output_too_large"


def test_invalid_embedded_base64_fails_closed_on_read():
    container = _valid_container()
    container["xlsx"]["base64"] = "not base64!"

    # The workbook payload is validated at parse time, so no consumer can reach
    # `decode_workbook` with undecodable bytes.
    with pytest.raises(RosterContainerInvalidError, match="strict base64"):
        parse_roster_container(_encode(container))


@pytest.mark.parametrize(
    ("stored", "expected"),
    [
        ("nurse-scheduling-20260201T000000Z.xlsx", "nurse-scheduling-20260201T000000Z.xlsx"),
        ("../../etc/passwd", "passwd"),
        ("C:\\Windows\\evil.xlsx", "evil.xlsx"),
        ('quote"; filename="other.xlsx', "quote___filename__other.xlsx"),
        ("line\r\nbreak.xlsx", "line__break.xlsx"),
        ("...", "schedule.xlsx"),
        ("", "schedule.xlsx"),
        (None, "schedule.xlsx"),
    ],
)
def test_download_filename_is_synthesized_safely(stored, expected):
    assert workbook_download_name({"xlsx": {"name": stored}}) == expected


def test_download_filename_is_length_bounded():
    name = workbook_download_name({"xlsx": {"name": "a" * 500 + ".xlsx"}})

    assert len(name) == 128
    assert set(name) == {"a"}


@pytest.mark.parametrize(
    "stored",
    [None, 7, "", "not a media type", "text/html", "text/html\r\nX-Injected: 1"],
)
def test_media_type_never_echoes_a_non_workbook_value(stored):
    assert workbook_media_type({"xlsx": {"mime": stored}}) == XLSX_MEDIA_TYPE


def test_media_type_accepts_the_stored_workbook_type():
    assert workbook_media_type({"xlsx": {"mime": XLSX_MEDIA_TYPE}}) == XLSX_MEDIA_TYPE
