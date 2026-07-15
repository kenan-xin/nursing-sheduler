"""Shared test helper for XLSX golden regression tests across solver backends."""

# This file is part of Nurse Scheduling Project, see <https://github.com/j3soon/nurse-scheduling>.
#
# Copyright (C) 2023-2026 Johnson Sun
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

# This test is mostly AI generated.

import logging
import os
import sys
from io import BytesIO

# Add the project root to the Python path so imports will work when running directly
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import nurse_scheduling
import nurse_scheduling.exporter as exporter
import pytest
from openpyxl import load_workbook
from pydantic import ValidationError

from .schedule_test_helper import CONTINUE_ON_ERROR, IGNORE_TESTS, TESTCASES_DIR, get_regression_testcases

WRITE_XLSX_GOLDEN = os.getenv("WRITE_XLSX_GOLDEN") == "1"


def _normalize_color(color):
    if color is None:
        return None
    return {
        "type": color.type,
        "rgb": color.rgb,
        "indexed": color.indexed,
        "theme": color.theme,
        "tint": color.tint,
    }


def _normalize_border_side(side):
    if side is None:
        return None
    return {
        "style": side.style,
        "color": _normalize_color(side.color),
    }


def _normalize_worksheet(ws):
    normalized_cells = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        normalized_row = []
        for cell in row:
            normalized_row.append(
                {
                    "value": cell.value,
                    "number_format": cell.number_format,
                    "font": {
                        "name": cell.font.name,
                        "size": cell.font.size,
                        "bold": cell.font.bold,
                        "italic": cell.font.italic,
                        "underline": cell.font.underline,
                        "color": _normalize_color(cell.font.color),
                    },
                    "fill": {
                        "fill_type": cell.fill.fill_type,
                        "fg": _normalize_color(cell.fill.fgColor),
                        "bg": _normalize_color(cell.fill.bgColor),
                    },
                    "alignment": {
                        "horizontal": cell.alignment.horizontal,
                        "vertical": cell.alignment.vertical,
                        "wrap_text": cell.alignment.wrap_text,
                    },
                    "border": {
                        "left": _normalize_border_side(cell.border.left),
                        "right": _normalize_border_side(cell.border.right),
                        "top": _normalize_border_side(cell.border.top),
                        "bottom": _normalize_border_side(cell.border.bottom),
                    },
                    "comment": None
                    if cell.comment is None
                    else {
                        "text": cell.comment.text,
                        "author": cell.comment.author,
                    },
                }
            )
        normalized_cells.append(normalized_row)
    return {
        "title": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "freeze_panes": ws.freeze_panes,
        "cells": normalized_cells,
    }


def _assert_workbook_matches_golden(generated_xlsx: bytes, golden_xlsx_path: str):
    generated_wb = load_workbook(BytesIO(generated_xlsx))
    golden_wb = load_workbook(golden_xlsx_path)

    generated_sheet_names = generated_wb.sheetnames
    golden_sheet_names = golden_wb.sheetnames
    assert generated_sheet_names == golden_sheet_names

    for sheet_name in generated_sheet_names:
        generated_ws = generated_wb[sheet_name]
        golden_ws = golden_wb[sheet_name]
        assert _normalize_worksheet(generated_ws) == _normalize_worksheet(golden_ws)


def _get_golden_xlsx_path(filepath: str, prettify: bool) -> str:
    if prettify:
        return f"{filepath[:-5]}.prettify.xlsx"
    return f"{filepath[:-5]}.xlsx"


def run_export_xlsx_regression_test(prettify: bool) -> None:
    tests = get_regression_testcases()
    total_tests = len(tests)
    error_count = 0

    for filepath in tests:
        base_filepath = os.path.splitext(os.path.basename(filepath))[0]
        test_dir = os.path.dirname(filepath)
        if base_filepath in IGNORE_TESTS:
            continue
        logging.info(
            "[prettify=%s] Testing XLSX '%s' ...",
            prettify,
            filepath[len(TESTCASES_DIR) + 1 :],
        )

        with open(filepath, "rb") as f:
            file_content = f.read()

        # If test should fail, preserve parity with schedule regression behavior.
        if os.path.isfile(f"{test_dir}/{base_filepath}.txt"):
            with open(f"{test_dir}/{base_filepath}.txt", "r", encoding="utf-8") as f:
                expected_err = f.read()
            with pytest.raises((ValidationError, ValueError)) as exc_info:
                nurse_scheduling.schedule(file_content, prettify=prettify)
            logging.info(f"Expected error: {expected_err.strip()}")
            logging.info(f"Actual error: {str(exc_info.value)}")
            assert expected_err.strip() in str(exc_info.value), (
                f"Expected error '{expected_err.strip()}' not found in actual error: {str(exc_info.value)}"
            )
            continue

        try:
            df, _solution, _score, _status, cell_export_info = nurse_scheduling.schedule(
                file_content,
                prettify=prettify,
            )
            if df is None:
                # Infeasible/no-solution scenarios have no table to export as XLSX.
                continue
            output = BytesIO()
            exporter.export_to_excel(df, output, cell_export_info)
            golden_xlsx_path = _get_golden_xlsx_path(filepath, prettify)

            if WRITE_XLSX_GOLDEN:
                with open(golden_xlsx_path, "wb") as f:
                    f.write(output.getvalue())
                continue

            if not os.path.isfile(golden_xlsx_path):
                raise FileNotFoundError(
                    f"Missing golden XLSX: {golden_xlsx_path}. "
                    "Run with WRITE_XLSX_GOLDEN=1 to generate/update golden files."
                )

            _assert_workbook_matches_golden(output.getvalue(), golden_xlsx_path)
        except ValidationError as e:
            logging.debug(f"Validation error for '{base_filepath}': {e}")
            error_count += 1
            if not CONTINUE_ON_ERROR:
                pytest.fail(f"Validation error for '{base_filepath}'")
            continue
        except Exception as e:
            logging.debug(f"Unexpected error for '{base_filepath}': {e}")
            error_count += 1
            if not CONTINUE_ON_ERROR:
                pytest.fail(f"Unexpected error for '{base_filepath}'")
            continue

    if error_count > 0:
        pytest.fail(f"Found {error_count}/{total_tests} errors during XLSX export testing")
    else:
        logging.info("All %s tests passed for prettify=%s", total_tests, prettify)
