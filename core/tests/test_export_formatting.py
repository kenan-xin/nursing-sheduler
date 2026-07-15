"""Tests for export formatting rules in XLSX export."""

# This file is part of Nurse Scheduling Project, see <https://github.com/j3soon/nurse-scheduling>.
#
# Copyright (C) 2023-2026 Johnson Sun
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

# This test is mostly AI generated.

import os
import sys
from io import BytesIO

import pytest
from openpyxl import load_workbook

# Add the project root to the Python path so imports will work when running directly
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from nurse_scheduling import exporter, schedule


def test_export_formatting_rules_apply_to_rows_columns_headers_and_cells():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-03
people:
  items:
    - id: n1
    - id: n2
    - id: n3
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 1
    qualifiedPeople: ALL
    date: ALL
    weight: -1
  - type: shift request
    person: n3
    date: ["2025-01-01", "2025-01-02", "2025-01-03"]
    shiftType: D
    weight: 100
export:
  formatting:
    - type: people header
      people: [n2]
      backgroundColor: "#f97316"
    - type: row
      people: [n2]
      backgroundColor: "#06b6d4"
      bottomBorderColor: "#ef4444"
      fontColor: "#dc2626"
    - type: cell
      people: [ALL]
      dates: [ALL]
      shiftTypes: [D]
      backgroundColor: "#1f2937"
    - type: date header
      dates: ["2025-01-01"]
      backgroundColor: "#a855f7"
    - type: column
      dates: ["2025-01-02"]
      backgroundColor: "#84cc16"
      bottomBorderColor: "#3b82f6"
      rightBorderColor: "#9ca3af"
"""

    df, _solution, _score, _status, cell_export_info = schedule(yaml_content, prettify=False)
    output = BytesIO()
    exporter.export_to_excel(df, output, cell_export_info)

    if os.getenv("WRITE_XLSX_ARTIFACT") == "1":
        artifact_path = os.path.join(
            os.path.dirname(__file__),
            "artifacts",
            "test_export_formatting.xlsx",
        )
        os.makedirs(os.path.dirname(artifact_path), exist_ok=True)
        with open(artifact_path, "wb") as f:
            f.write(output.getvalue())
        print(f"Wrote XLSX artifact: {artifact_path}")

    wb = load_workbook(output)
    ws = wb.active

    # Row target is n2 (Excel row 4). Row rule is after people-header rule, so it wins.
    assert ws["A4"].fill.fgColor.rgb == "FF06B6D4"
    assert ws["B4"].fill.fgColor.rgb == "FF06B6D4"
    assert ws["D4"].fill.fgColor.rgb == "FF06B6D4"
    assert ws["A4"].font.color.rgb == "FFDC2626"
    assert ws["B4"].font.color.rgb == "FFDC2626"
    assert ws["D4"].font.color.rgb == "FFDC2626"

    # Date-header target is date 2025-01-01 (Excel column B), first row only.
    assert ws["B1"].fill.fgColor.rgb == "FFA855F7"

    # Column target is 2025-01-02 (Excel column C), and should style the entire column.
    assert ws["C1"].fill.fgColor.rgb == "FF84CC16"
    assert ws["C2"].fill.fgColor.rgb == "FF84CC16"
    assert ws["C3"].fill.fgColor.rgb == "FF84CC16"
    assert ws["C4"].fill.fgColor.rgb == "FF84CC16"
    assert ws["C5"].fill.fgColor.rgb == "FF84CC16"
    # Score/Status summary rows should not be affected by full-column styling.
    assert ws["C6"].fill.fgColor.rgb == "00000000"
    assert ws["C7"].fill.fgColor.rgb == "00000000"

    # Cell rule for D applies to assigned schedule cells that are not overridden by column style.
    assert ws["B5"].fill.fgColor.rgb == "FF1F2937"
    assert ws["D5"].fill.fgColor.rgb == "FF1F2937"

    # A4 bottom border from row rule.
    assert ws["A4"].border.bottom.color is not None
    assert ws["A4"].border.bottom.color.rgb == "FFEF4444"
    # C4 bottom border from column rule overriding row rule.
    assert ws["C4"].border.bottom.color is not None
    assert ws["C4"].border.bottom.color.rgb == "FF3B82F6"
    assert ws["C4"].border.right.color is not None
    assert ws["C4"].border.right.color.rgb == "FF9CA3AF"


def test_export_formatting_rule_applies_to_history_cells():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-01
people:
  items:
    - id: n1
      history: [D]
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 0
export:
  formatting:
    - type: history
      people: [ALL]
      backgroundColor: "#fefce8"
"""

    df, _solution, _score, _status, cell_export_info = schedule(yaml_content, prettify=True)
    output = BytesIO()
    exporter.export_to_excel(df, output, cell_export_info)

    wb = load_workbook(output)
    ws = wb.active

    # With prettify enabled, the first history column is between the name column and dates.
    assert ws["B1"].value == "H-1"
    assert ws["B1"].fill.fgColor.rgb == "00000000"
    assert ws["B3"].fill.fgColor.rgb == "FFFEFCE8"


def test_export_formatting_cell_rule_applies_to_off_assignments():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-01
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 0
export:
  formatting:
    - type: cell
      people: [ALL]
      dates: [ALL]
      shiftTypes: [OFF]
      backgroundColor: "#22c55e"
"""

    df, _solution, _score, _status, cell_export_info = schedule(yaml_content, prettify=False)
    output = BytesIO()
    exporter.export_to_excel(df, output, cell_export_info)

    wb = load_workbook(output)
    ws = wb.active

    assert ws["B3"].fill.fgColor.rgb == "FF22C55E"


def test_export_formatting_rule_applies_to_history_headers():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-01
people:
  items:
    - id: n1
      history: [D]
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 0
export:
  formatting:
    - type: history header
      backgroundColor: "#fefce8"
"""

    df, _solution, _score, _status, cell_export_info = schedule(yaml_content, prettify=True)
    output = BytesIO()
    exporter.export_to_excel(df, output, cell_export_info)

    wb = load_workbook(output)
    ws = wb.active

    assert ws["B1"].value == "H-1"
    assert ws["B1"].fill.fgColor.rgb == "FFFEFCE8"
    assert ws["B3"].fill.fgColor.rgb == "00000000"


@pytest.mark.parametrize(
    ("formatting_yaml", "expected_message"),
    [
        (
            b"""
    - type: row
      people: [stale_person]
      backgroundColor: "#111111"
""",
            "Invalid person identifier 'stale_person'",
        ),
        (
            b"""
    - type: column
      dates: ["2025-01-02"]
      backgroundColor: "#111111"
""",
            "out of the range of start date and end date",
        ),
        (
            b"""
    - type: cell
      people: [n1]
      dates: ["2025-01-01"]
      shiftTypes: [stale_shift]
      backgroundColor: "#111111"
""",
            "Invalid shift type identifier 'stale_shift'",
        ),
    ],
)
def test_export_formatting_rejects_stale_references(formatting_yaml, expected_message):
    yaml_content = (
        b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-01
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 0
export:
  formatting:
"""
        + formatting_yaml
    )

    with pytest.raises(ValueError, match=expected_message):
        schedule(yaml_content, prettify=True)


@pytest.mark.parametrize(
    ("extra_layout_yaml", "expected_message"),
    [
        (
            b"""
  extraColumns:
    - type: count
      header: Stale date
      countDates: ["2025-01-02"]
      countShiftTypes: [D]
""",
            "out of the range of start date and end date",
        ),
        (
            b"""
  extraColumns:
    - type: count
      header: Stale shift
      countDates: ["2025-01-01"]
      countShiftTypes: [stale_shift]
""",
            "Unknown shift type ID: stale_shift",
        ),
        (
            b"""
  extraColumns:
    - type: count
      header: Stale coefficient shift
      countDates: ["2025-01-01"]
      countShiftTypes: [D]
      countShiftTypeCoefficients:
        - [stale_shift, 2]
""",
            "Unknown shift type ID: stale_shift",
        ),
        (
            b"""
  extraColumns:
    - type: count
      header: Uncovered coefficient shift
      countDates: ["2025-01-01"]
      countShiftTypes: [D]
      countShiftTypeCoefficients:
        - [A, 2]
""",
            "must be covered by countShiftTypes",
        ),
        (
            b"""
  extraColumns:
    - type: count
      header: Invalid coefficient
      countDates: ["2025-01-01"]
      countShiftTypes: [D]
      countShiftTypeCoefficients:
        - [D, 0]
""",
            "must be at least 1",
        ),
        (
            b"""
  extraColumns:
    - type: count
      header: Fractional coefficient
      countDates: ["2025-01-01"]
      countShiftTypes: [D]
      countShiftTypeCoefficients:
        - [D, 1.5]
""",
            "Input should be a valid integer",
        ),
        (
            b"""
  extraColumns:
    - type: count
      header: Duplicate coefficient
      countDates: ["2025-01-01"]
      countShiftTypes: [D]
      countShiftTypeCoefficients:
        - [D, 2]
        - [D, 3]
""",
            "Duplicate export extra column coefficient",
        ),
        (
            b"""
  extraRows:
    - type: count
      header: Stale person
      countPeople: [stale_person]
      countShiftTypes: [D]
""",
            "Unknown person ID: stale_person",
        ),
        (
            b"""
  extraRows:
    - type: count
      header: Stale row shift
      countPeople: [n1]
      countShiftTypes: [stale_shift]
""",
            "Unknown shift type ID: stale_shift",
        ),
    ],
)
def test_export_extra_layout_rejects_stale_references(extra_layout_yaml, expected_message):
    yaml_content = (
        b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-01
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
    - id: A
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: D
    requiredNumPeople: 0
export:
  formatting: []
"""
        + extra_layout_yaml
    )

    with pytest.raises(ValueError, match=expected_message):
        schedule(yaml_content, prettify=True)


def test_export_extra_column_rejects_overlapping_expanded_coefficients():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-01
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
    - id: A
  groups:
    - id: WORK
      members: [D, A]
preferences:
  - type: at most one shift per day
export:
  extraColumns:
    - type: count
      header: Invalid overlap
      countDates: [ALL]
      countShiftTypes: [D, WORK]
      countShiftTypeCoefficients:
        - [D, 2]
        - [WORK, 3]
"""

    with pytest.raises(ValueError, match="Duplicate export extra column coefficient"):
        schedule(yaml_content, prettify=True)


def test_export_extra_column_rejects_overlapping_explicit_coefficient_one():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-01
people:
  items:
    - id: n1
shiftTypes:
  items:
    - id: D
    - id: A
  groups:
    - id: WORK
      members: [D, A]
preferences:
  - type: at most one shift per day
export:
  extraColumns:
    - type: count
      header: Invalid overlap
      countDates: [ALL]
      countShiftTypes: [D, WORK]
      countShiftTypeCoefficients:
        - [D, 1]
        - [WORK, 2]
"""

    with pytest.raises(ValueError, match="Duplicate export extra column coefficient"):
        schedule(yaml_content, prettify=True)


def test_export_xlsx_handles_unequal_trimmed_history_columns():
    yaml_content = b"""
apiVersion: alpha
dates:
  range:
    startDate: 2025-01-01
    endDate: 2025-01-01
people:
  items:
    - id: n1
      history: [N]
    - id: n2
      history: [A, D, N]
shiftTypes:
  items:
    - id: A
    - id: D
    - id: N
preferences:
  - type: at most one shift per day
  - type: shift type requirement
    shiftType: N
    requiredNumPeople: 0
export:
  formatting:
    - type: history
      people: [ALL]
      backgroundColor: "#fefce8"
    - type: history header
      backgroundColor: "#bfdbfe"
"""

    df, _solution, _score, _status, cell_export_info = schedule(yaml_content, prettify=True)
    output = BytesIO()
    exporter.export_to_excel(df, output, cell_export_info)

    wb = load_workbook(output)
    ws = wb.active

    assert [ws["B1"].value, ws["C1"].value, ws["D1"].value] == ["H-3", "H-2", "H-1"]
    assert [ws["B3"].value, ws["C3"].value, ws["D3"].value] == [None, None, "N"]
    assert [ws["B4"].value, ws["C4"].value, ws["D4"].value] == ["A", "D", "N"]
    assert ws["B1"].fill.fgColor.rgb == "FFBFDBFE"
    assert ws["D3"].fill.fgColor.rgb == "FFFEFCE8"
