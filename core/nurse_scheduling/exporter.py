"""Schedule export helpers for CSV/XLSX and rendered outputs."""

# This file is part of Nurse Scheduling Project, see <https://github.com/j3soon/nurse-scheduling>.
#
# Copyright (C) 2023-2026 Johnson Sun
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

from io import BytesIO, StringIO
from copy import copy

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles.borders import Side

from .context import Context
from . import utils, models, constants


def _get_font_color_for_background(hex_color: str) -> str:
    """Return ARGB font color (black/white) for readable contrast on a hex background."""
    r = int(hex_color[1:3], 16)
    g = int(hex_color[3:5], 16)
    b = int(hex_color[5:7], 16)
    luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255
    # Match frontend threshold used in getPickerDisplay().
    return "FF000000" if luminance > 0.6 else "FFFFFFFF"


def _build_custom_export_style_info(
    ctx: Context,
    n_rows: int,
    n_cols: int,
    n_leading_rows: int,
    n_leading_cols: int,
    n_history_cols: int,
):
    """Build cell-level style overrides from ctx.export.formatting."""
    if not ctx.export or not ctx.export.formatting:
        return {}

    style_map = {}

    def set_style(
        row_idx: int,
        col_idx: int,
        background_color: str | None,
        bottom_border_color: str | None,
        right_border_color: str | None,
        font_color: str | None,
    ):
        if row_idx < 0 or row_idx >= n_rows or col_idx < 0 or col_idx >= n_cols:
            return
        key = (row_idx + 1, col_idx + 1)  # Store in 1-based Excel coordinates
        if key not in style_map:
            style_map[key] = {}
        if background_color:
            style_map[key]["backgroundColor"] = background_color
        if bottom_border_color:
            style_map[key]["bottomBorderColor"] = bottom_border_color
        if right_border_color:
            style_map[key]["rightBorderColor"] = right_border_color
        if font_color:
            style_map[key]["fontColor"] = font_color

    for rule in ctx.export.formatting:
        _validate_export_formatting_rule_usage(rule)

        target_people = set()
        target_dates = set()
        target_shift_types = set()

        if rule.type in ("row", "people header", "history", "cell"):
            for target in rule.people:
                if target not in ctx.map_pid_p:
                    raise ValueError(
                        f"Invalid person identifier '{target}' in export formatting rule with type '{rule.type}'"
                    )
                target_people.update(ctx.map_pid_p[target])

        if rule.type in ("column", "date header", "cell"):
            for target in rule.dates:
                target_dates.update(utils.parse_dates(target, ctx.map_did_d, ctx.dates.range))

        if rule.type == "cell":
            for target in rule.shiftTypes:
                if target not in ctx.map_sid_s:
                    raise ValueError(
                        f"Invalid shift type identifier '{target}' in export formatting rule with type 'cell'"
                    )
                target_shift_types.update(ctx.map_sid_s[target])

        if rule.type == "row":
            for p in target_people:
                row_idx = n_leading_rows + p
                for col_idx in range(n_cols):
                    set_style(
                        row_idx,
                        col_idx,
                        rule.backgroundColor,
                        rule.bottomBorderColor,
                        rule.rightBorderColor,
                        rule.fontColor,
                    )

        elif rule.type == "people header":
            for p in target_people:
                row_idx = n_leading_rows + p
                set_style(
                    row_idx,
                    0,
                    rule.backgroundColor,
                    rule.bottomBorderColor,
                    rule.rightBorderColor,
                    rule.fontColor,
                )

        elif rule.type == "column":
            score_row_idx = n_leading_rows + len(ctx.people.items)
            status_row_idx = score_row_idx + 1
            for d in target_dates:
                col_idx = n_leading_cols + n_history_cols + d
                for row_idx in range(n_rows):
                    if row_idx in (score_row_idx, status_row_idx):
                        # Skip styling for score/status summary rows since they are not part of the main schedule grid and should not be affected by column styles.
                        continue
                    set_style(
                        row_idx,
                        col_idx,
                        rule.backgroundColor,
                        rule.bottomBorderColor,
                        rule.rightBorderColor,
                        rule.fontColor,
                    )

        elif rule.type == "date header":
            for d in target_dates:
                col_idx = n_leading_cols + n_history_cols + d
                set_style(
                    0,
                    col_idx,
                    rule.backgroundColor,
                    rule.bottomBorderColor,
                    rule.rightBorderColor,
                    rule.fontColor,
                )

        elif rule.type == "history header":
            for col_idx in range(n_leading_cols, n_leading_cols + n_history_cols):
                set_style(
                    0,
                    col_idx,
                    rule.backgroundColor,
                    rule.bottomBorderColor,
                    rule.rightBorderColor,
                    rule.fontColor,
                )

        elif rule.type == "cell":
            if rule.when:
                for d, p, _pref, _requested_shift_type in _iter_matching_cell_preferences(
                    ctx,
                    target_people=target_people,
                    target_dates=target_dates,
                    target_shift_types=target_shift_types,
                    condition=rule.when,
                ):
                    row_idx = n_leading_rows + p
                    col_idx = n_leading_cols + n_history_cols + d
                    set_style(
                        row_idx,
                        col_idx,
                        rule.backgroundColor,
                        rule.bottomBorderColor,
                        rule.rightBorderColor,
                        rule.fontColor,
                    )
            else:
                actual_target_shift_types = {
                    s
                    for s in target_shift_types
                    if s in (constants.OFF_sid, constants.LEAVE_sid) or 0 <= s < ctx.n_shift_types
                }
                for d in target_dates:
                    for p in target_people:
                        if (d, p) not in ctx.map_dp_s:
                            continue
                        assigned_shift_types = [
                            s for s in ctx.map_dp_s[(d, p)] if ctx.solver.get_value(ctx.shifts[(d, s, p)]) == 1
                        ]
                        if ctx.solver.get_value(ctx.offs[(d, p)]) == 1:
                            assigned_shift_types.append(constants.OFF_sid)
                        if ctx.solver.get_value(ctx.leaves[(d, p)]) == 1:
                            assigned_shift_types.append(constants.LEAVE_sid)
                        if not any(s in actual_target_shift_types for s in assigned_shift_types):
                            continue
                        row_idx = n_leading_rows + p
                        col_idx = n_leading_cols + n_history_cols + d
                        set_style(
                            row_idx,
                            col_idx,
                            rule.backgroundColor,
                            rule.bottomBorderColor,
                            rule.rightBorderColor,
                            rule.fontColor,
                        )

        elif rule.type == "history":
            for p in target_people:
                row_idx = n_leading_rows + p
                for col_idx in range(n_leading_cols, n_leading_cols + n_history_cols):
                    set_style(
                        row_idx,
                        col_idx,
                        rule.backgroundColor,
                        rule.bottomBorderColor,
                        rule.rightBorderColor,
                        rule.fontColor,
                    )

    return style_map


def _parse_extra_column_coefficients(ctx: Context, rule, count_shift_types: list[int]) -> dict[int, int]:
    coefficients = dict.fromkeys(count_shift_types, 1)
    coefficient_entries = rule.countShiftTypeCoefficients or []
    selected_sids = set(count_shift_types)
    coefficient_sids = set()

    for shift_type_id, coefficient in coefficient_entries:
        if coefficient < 1:
            raise ValueError(f"Export extra column coefficient for '{shift_type_id}' must be at least 1.")

        expanded_sids = utils.parse_sids(shift_type_id, ctx.map_sid_s)
        if not set(expanded_sids).issubset(selected_sids):
            raise ValueError(
                f"Export extra column coefficient for '{shift_type_id}' must be covered by countShiftTypes."
            )
        duplicate_sids = coefficient_sids.intersection(expanded_sids)
        if duplicate_sids:
            raise ValueError(f"Duplicate export extra column coefficient for '{shift_type_id}'.")
        coefficient_sids.update(expanded_sids)

        for s in expanded_sids:
            coefficients[s] = coefficient

    return coefficients


def _count_extra_column_for_person(ctx: Context, p: int, count_dates, count_shift_types, coefficients) -> int:
    count = 0
    for d in count_dates:
        if constants.OFF_sid in count_shift_types and ctx.solver.get_value(ctx.offs[(d, p)]) == 1:
            count += coefficients[constants.OFF_sid]
            continue
        if constants.LEAVE_sid in count_shift_types and ctx.solver.get_value(ctx.leaves[(d, p)]) == 1:
            count += coefficients[constants.LEAVE_sid]
            continue
        count += sum(
            coefficients[s]
            for s in count_shift_types
            if 0 <= s < ctx.n_shift_types and ctx.solver.get_value(ctx.shifts[(d, s, p)]) == 1
        )
    return count


def _count_extra_row_for_date(ctx: Context, d: int, count_people, count_shift_types) -> int:
    count = 0
    for p in count_people:
        if constants.OFF_sid in count_shift_types and ctx.solver.get_value(ctx.offs[(d, p)]) == 1:
            count += 1
            continue
        if constants.LEAVE_sid in count_shift_types and ctx.solver.get_value(ctx.leaves[(d, p)]) == 1:
            count += 1
            continue
        if any(
            0 <= s < ctx.n_shift_types and ctx.solver.get_value(ctx.shifts[(d, s, p)]) == 1 for s in count_shift_types
        ):
            count += 1
    return count


def _validate_export_formatting_rule_usage(rule):
    if rule.type != "cell" and getattr(rule, "when", None):
        raise ValueError("export formatting 'when' is only supported for rules with type 'cell'")
    if rule.type != "cell" and (getattr(rule, "appendText", None) or getattr(rule, "note", None)):
        raise ValueError("export formatting annotations are only supported for rules with type 'cell'")


def _get_shift_request_shape(ctx: Context, person_target, date_target) -> str:
    person_id = person_target
    person_item_ids = {person.id for person in ctx.people.items}
    people_group_ids = {group.id for group in ctx.people.groups}
    date_item_ids = {str(date) for date in ctx.dates.items}
    date_group_ids = {group.id for group in ctx.dates.groups}
    date_keyword_ids = set(constants.MAP_DATE_KEYWORD_TO_FILTER) | set(constants.MAP_WEEKDAY_TO_STR)
    date_id = str(date_target)

    if person_id in person_item_ids:
        person_shape = "person-item"
    elif person_id in people_group_ids:
        person_shape = "people-group"
    else:
        return "unknown"

    if date_id in date_item_ids:
        date_shape = "date-item"
    elif date_id in date_group_ids or date_id in date_keyword_ids:
        date_shape = "date-group"
    else:
        try:
            parsed_dates = utils.parse_dates(date_id, ctx.map_did_d, ctx.dates.range)
        except ValueError:
            return "unknown"
        date_shape = "date-item" if len(parsed_dates) == 1 else "date-group"

    return f"{person_shape}-to-{date_shape}"


def _render_export_template(template: str, *, pref, requested_shift_type: str, total_abs_weight: int | float) -> str:
    return (
        template.replace("{shiftType}", requested_shift_type)
        .replace("{weight}", str(pref.weight))
        .replace("{absWeight}", str(abs(pref.weight)))
        .replace("{totalAbsWeight}", str(total_abs_weight))
    )


def _format_requested_shift_type(shift_type_targets) -> str:
    return ", ".join(str(target) for target in shift_type_targets)


def _build_cell_annotation_rules(ctx: Context):
    if not ctx.export or not ctx.export.formatting:
        return []

    annotation_rules = []
    for rule in ctx.export.formatting:
        _validate_export_formatting_rule_usage(rule)

        if rule.type != "cell" or not rule.when or (not rule.appendText and not rule.note):
            continue

        target_people = set()
        target_dates = set()
        target_shift_types = set()
        for target in rule.people:
            if target not in ctx.map_pid_p:
                raise ValueError(
                    f"Invalid person identifier '{target}' in export formatting rule with type '{rule.type}'"
                )
            target_people.update(ctx.map_pid_p[target])
        for target in rule.dates:
            target_dates.update(utils.parse_dates(target, ctx.map_did_d, ctx.dates.range))
        for target in rule.shiftTypes:
            if target not in ctx.map_sid_s:
                raise ValueError(f"Invalid shift type identifier '{target}' in export formatting rule with type 'cell'")
            target_shift_types.update(ctx.map_sid_s[target])
        annotation_rules.append(
            {
                "rule": rule,
                "people": target_people,
                "dates": target_dates,
                "shift_types": target_shift_types,
            }
        )

    return annotation_rules


def _export_preference_condition_matches(ctx: Context, condition, pref, *, request_shape: str, satisfied: bool):
    pref_condition = condition.preference
    unsupported_types = set(pref_condition.types) - {models.SHIFT_REQUEST}
    if unsupported_types:
        raise ValueError(f"Unsupported export formatting preference condition type(s): {sorted(unsupported_types)}")
    if pref.type not in pref_condition.types:
        return False
    if pref_condition.satisfied is not None and pref_condition.satisfied != satisfied:
        return False
    if pref_condition.weightRange is not None:
        if len(pref_condition.weightRange) != 2:
            raise ValueError("export formatting preference weightRange must contain exactly two values")
        min_weight, max_weight = pref_condition.weightRange
        if min_weight > max_weight:
            raise ValueError("export formatting preference weightRange minimum must be less than or equal to maximum")
        if pref.weight < min_weight or pref.weight > max_weight:
            return False
    if pref_condition.requestShape is not None and constants.ALL not in pref_condition.requestShape:
        if request_shape not in pref_condition.requestShape:
            return False
    return True


def _iter_expanded_shift_request_targets(ctx: Context, pref):
    """Expand compact frontend shift request date targets for export matching.

    Frontend edits compact real date items together, while date groups remain
    separate preferences so overlapping groups can stack. Older saved YAML can
    still contain mixed targets in ``pref.date``. Each entry is treated as a
    distinct matrix target: either an individual date column or a date-group
    column. Shape matching must use that original target, plus the person and
    shift-type target shapes, before expanding to concrete schedule dates for
    the exported sheet.
    """
    person_targets = utils.ensure_list(pref.person)
    date_targets = utils.ensure_list(pref.date)
    shift_type_targets = utils.ensure_list(pref.shiftType)
    if len(person_targets) != 1 or len(shift_type_targets) != 1:
        for date_target in date_targets:
            yield date_target, utils.parse_dates(date_target, ctx.map_did_d, ctx.dates.range), "unknown"
        return

    person_target = person_targets[0]
    for date_target in date_targets:
        yield (
            date_target,
            utils.parse_dates(date_target, ctx.map_did_d, ctx.dates.range),
            _get_shift_request_shape(ctx, person_target, date_target),
        )


def _is_shift_request_satisfied(ctx: Context, pref, *, d: int, p: int, shift_types: list[int]) -> bool:
    """Return whether a cell-level shift request is satisfied by the solved schedule."""

    def _state_var(s):
        if s == constants.OFF_sid:
            return ctx.offs[(d, p)]
        if s == constants.LEAVE_sid:
            return ctx.leaves[(d, p)]
        return ctx.shifts[(d, s, p)]

    requested_state_is_assigned = any(ctx.solver.get_value(_state_var(s)) == 1 for s in shift_types)
    return requested_state_is_assigned if pref.weight > 0 else not requested_state_is_assigned


def _iter_matching_cell_preferences(
    ctx: Context,
    *,
    target_people: set[int],
    target_dates: set[int],
    target_shift_types: set[int],
    condition,
):
    for pref in ctx.preferences:
        if pref.type != models.SHIFT_REQUEST:
            continue
        if pref.weight == 0:
            continue

        shift_type_targets = utils.ensure_list(pref.shiftType)
        ss = utils.parse_sids(shift_type_targets, ctx.map_sid_s)
        ps = utils.parse_pids(pref.person, ctx.map_pid_p)
        if not any(s in target_shift_types for s in ss):
            continue

        requested_shift_type = _format_requested_shift_type(shift_type_targets)

        for _date_target, ds, request_shape in _iter_expanded_shift_request_targets(ctx, pref):
            for d in ds:
                if d not in target_dates:
                    continue
                for p in ps:
                    if p not in target_people:
                        continue
                    satisfied = _is_shift_request_satisfied(ctx, pref, d=d, p=p, shift_types=ss)
                    if _export_preference_condition_matches(
                        ctx, condition, pref, request_shape=request_shape, satisfied=satisfied
                    ):
                        yield d, p, pref, requested_shift_type


def get_people_versus_date_dataframe(ctx: Context, prettify: bool = False):
    # Initialize dataframe with size including leading rows and columns
    n_leading_rows, n_leading_cols = 2, 1
    n_trailing_rows, n_trailing_cols = 2, 0

    # Dictionary to track cells with generated Excel notes
    cell_comment_info = {}

    n_history_cols = 0
    # Add history columns after the name column (only if prettify is enabled)
    if prettify:
        max_history_length = max((len(person.history) for person in ctx.people.items if person.history), default=0)
        n_history_cols = max_history_length

    extra_column_rules = ctx.export.extraColumns if prettify and ctx.export else []
    extra_row_rules = ctx.export.extraRows if prettify and ctx.export else []
    # Add extra columns and rows for prettify mode
    extra_cols = (1 + len(extra_column_rules)) if extra_column_rules else 0  # Empty separator + configured columns
    extra_rows = (1 + len(extra_row_rules)) if extra_row_rules else 0  # Empty separator + configured rows

    df = pd.DataFrame(
        "",
        index=range(n_leading_rows + len(ctx.people.items) + n_trailing_rows + extra_rows),
        columns=range(n_leading_cols + n_history_cols + len(ctx.dates.items) + n_trailing_cols + extra_cols),
        # We could cast every write to str, but object dtype is much simpler for mixed cells.
        dtype=object,
    )

    # Fill history column headers (only if prettify is enabled)
    if n_history_cols > 0:
        # - row 0 contains history position labels (H-1, H-2, etc.)
        # - row 1 contains "History" label
        for h in range(n_history_cols):
            df.iloc[0, n_leading_cols + h] = f"H-{n_history_cols - h}"
            df.iloc[1, n_leading_cols + h] = "History"

    # Fill day numbers and weekdays
    # - row 0 contains day number
    # - row 1 contains weekday
    for d, date in enumerate(ctx.dates.items):
        col_idx = n_leading_cols + n_history_cols + d
        if ctx.dates.items[0].year != ctx.dates.items[-1].year:
            df.iloc[0, col_idx] = f"{date.year}/{date.month}/{date.day}"
        elif ctx.dates.items[0].month != ctx.dates.items[-1].month:
            df.iloc[0, col_idx] = f"{date.month}/{date.day}"
        else:
            df.iloc[0, col_idx] = date.day
        df.iloc[1, col_idx] = date.strftime("%a")

    # Fill person descriptions and history
    # - column 0 contains person description
    # - columns 1 to n_history_cols contain history data (padded with empty strings, only if prettify)
    for p, person in enumerate(ctx.people.items):
        df.iloc[n_leading_rows + p, 0] = person.id

        # Fill history columns with proper padding (only if prettify is enabled)
        if n_history_cols > 0:
            if person.history:
                history = person.history
                # Pad with empty strings at the front if history is shorter than n_history_cols
                padded_history = [""] * max(0, n_history_cols - len(history)) + history
                for h in range(n_history_cols):
                    df.iloc[n_leading_rows + p, n_leading_cols + h] = padded_history[h]
            else:
                # Fill with empty strings if no history
                for h in range(n_history_cols):
                    df.iloc[n_leading_rows + p, n_leading_cols + h] = ""

    annotation_rules = _build_cell_annotation_rules(ctx) if prettify else []
    cell_annotations = {}
    if annotation_rules:
        for annotation_rule_data in annotation_rules:
            rule = annotation_rule_data["rule"]
            matches_by_cell = {}
            for d, p, pref, requested_shift_type in _iter_matching_cell_preferences(
                ctx,
                target_people=annotation_rule_data["people"],
                target_dates=annotation_rule_data["dates"],
                target_shift_types=annotation_rule_data["shift_types"],
                condition=rule.when,
            ):
                matches_by_cell.setdefault((d, p), []).append((pref, requested_shift_type))

            for (d, p), matches in matches_by_cell.items():
                if (d, p) not in cell_annotations:
                    cell_annotations[(d, p)] = {"append_text": [], "notes": []}
                total_abs_weight = sum(abs(pref.weight) for pref, _requested_shift_type in matches)
                for pref, requested_shift_type in matches:
                    rendered_context = {
                        "pref": pref,
                        "requested_shift_type": requested_shift_type,
                        "total_abs_weight": total_abs_weight,
                    }
                    if rule.appendText:
                        cell_annotations[(d, p)]["append_text"].append(
                            _render_export_template(rule.appendText, **rendered_context)
                        )
                if rule.note:
                    pref, requested_shift_type = matches[0]
                    cell_annotations[(d, p)]["notes"].append(
                        _render_export_template(
                            rule.note.text,
                            pref=pref,
                            requested_shift_type=requested_shift_type,
                            total_abs_weight=total_abs_weight,
                        )
                    )

    # Set cell values based on solver results
    solver = ctx.solver

    for d, p in ctx.map_dp_s.keys():
        col_idx = n_leading_cols + n_history_cols + d
        assert df.iloc[n_leading_rows + p, col_idx] == ""
        cell_value = ""
        for s in ctx.map_dp_s[(d, p)]:
            if solver.get_value(ctx.shifts[(d, s, p)]) == 1:
                if cell_value != "":
                    cell_value += ", "
                cell_value += ctx.shiftTypes.items[s].id
        # Render a paid-leave day distinctly (never blank). OFF still renders
        # blank; leave is a worked-day peer that must be visible on the roster.
        if solver.get_value(ctx.leaves[(d, p)]) == 1:
            cell_value = "Leave"
        if prettify and (d, p) in cell_annotations:
            for append_text in cell_annotations[(d, p)]["append_text"]:
                cell_value += append_text
            if cell_annotations[(d, p)]["notes"]:
                excel_row = n_leading_rows + p + 1  # +1 for 1-based Excel indexing
                excel_col = n_leading_cols + n_history_cols + d + 1  # +1 for 1-based Excel indexing
                cell_comment_info[(excel_row, excel_col)] = cell_annotations[(d, p)]["notes"]
        df.iloc[n_leading_rows + p, col_idx] = cell_value

    # Fill objective value
    df.iloc[n_leading_rows + len(ctx.people.items), 0] = "Score"
    df.iloc[n_leading_rows + len(ctx.people.items), n_leading_cols + n_history_cols] = solver.get_objective_value()
    # Fill solver status
    df.iloc[n_leading_rows + len(ctx.people.items) + 1, 0] = "Status"
    df.iloc[n_leading_rows + len(ctx.people.items) + 1, n_leading_cols + n_history_cols] = ctx.solver_status

    # Sanity check the three day-states: OFF renders blank, LEAVE renders
    # "Leave", and a worked day renders a non-empty shift id.
    if not prettify:
        for d, p in ctx.offs.keys():
            col_idx = n_leading_cols + n_history_cols + d
            cell = df.iloc[n_leading_rows + p, col_idx]
            if solver.get_value(ctx.offs[(d, p)]) == 1:
                assert cell == ""
            elif solver.get_value(ctx.leaves[(d, p)]) == 1:
                assert cell == "Leave"
            else:
                assert cell != ""

    if prettify:
        extra_col_start = n_leading_cols + n_history_cols + len(ctx.dates.items) + 1
        for rule_idx, rule in enumerate(extra_column_rules):
            col_idx = extra_col_start + rule_idx
            count_dates = utils.parse_dates(rule.countDates, ctx.map_did_d, ctx.dates.range)
            count_shift_types = utils.parse_sids(rule.countShiftTypes, ctx.map_sid_s)
            coefficients = _parse_extra_column_coefficients(ctx, rule, count_shift_types)
            df.iloc[1, col_idx] = rule.header
            for p in range(len(ctx.people.items)):
                df.iloc[n_leading_rows + p, col_idx] = _count_extra_column_for_person(
                    ctx,
                    p,
                    count_dates,
                    count_shift_types,
                    coefficients,
                )

        extra_row_start = n_leading_rows + len(ctx.people.items) + n_trailing_rows + 1
        for rule_idx, rule in enumerate(extra_row_rules):
            row_idx = extra_row_start + rule_idx
            count_people = utils.parse_pids(rule.countPeople, ctx.map_pid_p)
            count_shift_types = utils.parse_sids(rule.countShiftTypes, ctx.map_sid_s)
            df.iloc[row_idx, 0] = rule.header
            for d in range(len(ctx.dates.items)):
                df.iloc[row_idx, n_leading_cols + n_history_cols + d] = _count_extra_row_for_date(
                    ctx,
                    d,
                    count_people,
                    count_shift_types,
                )

    # Apply default styling and borders if prettify is enabled
    if prettify:
        # Create a styler object to apply conditional formatting
        def apply_styling(df):
            # Create a style DataFrame with the same shape as the original
            style_df = pd.DataFrame("", index=df.index, columns=df.columns)

            # Apply center alignment to all cells
            for row_idx in range(len(df)):
                for col_idx in range(len(df.columns)):
                    style_df.iloc[row_idx, col_idx] = "text-align: center"

            # Add borders to separate regions
            # Horizontal borders
            header_row_end = n_leading_rows - 1  # End of header region
            people_row_end = header_row_end + len(ctx.people.items)  # End of people region
            summary_row_end = people_row_end + n_trailing_rows  # End of summary region
            extra_rows_end = summary_row_end + len(extra_row_rules) + 1

            # Vertical borders
            name_col_end = n_leading_cols - 1  # End of name column
            history_col_end = name_col_end + n_history_cols  # End of history columns
            date_col_end = history_col_end + len(ctx.dates.items)  # End of date columns
            extra_columns_end = date_col_end + len(extra_column_rules) + 1

            # Apply borders to all cells, then add specific border styles
            for row_idx in range(len(df)):
                for col_idx in range(len(df.columns)):
                    base_style = style_df.iloc[row_idx, col_idx]
                    borders = []

                    # Add horizontal borders
                    if row_idx in [
                        header_row_end,
                        people_row_end,
                        summary_row_end,
                        extra_rows_end,
                    ]:
                        borders.append("border-bottom: 2px solid #374151")

                    # Add vertical borders
                    if col_idx in [
                        name_col_end,
                        history_col_end,
                        date_col_end,
                        extra_columns_end,
                    ]:
                        borders.append("border-right: 2px solid #374151")

                    # Combine base style with borders
                    if borders:
                        border_style = "; ".join(borders)
                        if base_style:
                            style_df.iloc[row_idx, col_idx] = f"{base_style}; {border_style}"
                        else:
                            style_df.iloc[row_idx, col_idx] = border_style

            return style_df

        # Apply the styling and return the styled DataFrame
        styled_df = df.style.apply(lambda x: apply_styling(df), axis=None)
        style_info = _build_custom_export_style_info(
            ctx, len(df.index), len(df.columns), n_leading_rows, n_leading_cols, n_history_cols
        )
        for rule_idx, rule in enumerate(extra_column_rules):
            if rule.rightBorderColor:
                col_idx = extra_col_start + rule_idx
                for row_idx in range(len(df.index)):
                    style_info.setdefault((row_idx + 1, col_idx + 1), {})["rightBorderColor"] = rule.rightBorderColor
        for rule_idx, rule in enumerate(extra_row_rules):
            if rule.bottomBorderColor:
                row_idx = extra_row_start + rule_idx
                for col_idx in range(len(df.columns)):
                    style_info.setdefault((row_idx + 1, col_idx + 1), {})["bottomBorderColor"] = rule.bottomBorderColor
        return styled_df, {"comments": cell_comment_info, "styles": style_info}

    style_info = _build_custom_export_style_info(
        ctx, len(df.index), len(df.columns), n_leading_rows, n_leading_cols, n_history_cols
    )
    return df, {"comments": cell_comment_info, "styles": style_info}


def export_to_excel(df, output_buffer, cell_export_info=None):
    """
    Export DataFrame to Excel with frozen panes at B3 (first two rows and first column).
    Also applies configured cell notes and styles.

    Args:
        output_buffer: BytesIO buffer to write to
        cell_export_info: Dictionary containing comments/styles from dataframe export
    """

    # Write to a temporary BytesIO buffer first
    temp_buffer = BytesIO()
    df.to_excel(temp_buffer, index=False, header=False)
    temp_buffer.seek(0)

    # Load the workbook to apply additional formatting
    wb = load_workbook(temp_buffer)
    ws = wb.active

    # Freeze the first two rows and first column (B3 is the cell after frozen area)
    ws.freeze_panes = "B3"

    comment_info = {}
    style_info = {}
    if cell_export_info is not None:
        if not isinstance(cell_export_info, dict) or not {"comments", "styles"}.issuperset(cell_export_info):
            raise ValueError("cell_export_info must be a dictionary with optional 'comments' and 'styles' keys")
        comment_info = cell_export_info.get("comments") or {}
        style_info = cell_export_info.get("styles") or {}

    # Put notes in a separate sheet so many annotations remain readable without
    # overlapping native Excel comment popups.
    if comment_info:
        notes_ws = wb.create_sheet("Notes")
        notes_ws.append(["Cell", "Schedule Value", "Note"])
        notes_ws.freeze_panes = "A2"
        notes_ws.auto_filter.ref = "A1:C1"
        notes_ws.column_dimensions["A"].width = 14
        notes_ws.column_dimensions["B"].width = 24
        notes_ws.column_dimensions["C"].width = 80

        schedule_sheet_name = ws.title.replace("'", "''")
        notes_sheet_name = notes_ws.title.replace("'", "''")

        for (row, col), notes in comment_info.items():
            cell = ws.cell(row=row, column=col)
            if not all(isinstance(note, str) for note in notes):
                raise ValueError("cell_export_info comments must be lists of strings")
            first_note_row = notes_ws.max_row + 1
            for note in notes:
                notes_ws.append([cell.coordinate, cell.value, note])
                note_cell = notes_ws.cell(row=notes_ws.max_row, column=1)
                note_cell.hyperlink = f"#'{schedule_sheet_name}'!{cell.coordinate}"
                note_cell.style = "Hyperlink"
            if notes:
                cell.hyperlink = f"#'{notes_sheet_name}'!A{first_note_row}"

    # Apply custom export formatting styles.
    if style_info:
        for (row, col), styles in style_info.items():
            cell = ws.cell(row=row, column=col)

            background_color = styles.get("backgroundColor")
            if background_color:
                argb = f"FF{background_color[1:].upper()}"
                cell.fill = PatternFill(fill_type="solid", start_color=argb, end_color=argb)
                updated_font = copy(cell.font)
                updated_font.color = _get_font_color_for_background(background_color)
                cell.font = updated_font

            font_color = styles.get("fontColor")
            if font_color:
                updated_font = copy(cell.font)
                updated_font.color = f"FF{font_color[1:].upper()}"
                cell.font = updated_font

            bottom_border_color = styles.get("bottomBorderColor")
            right_border_color = styles.get("rightBorderColor")
            if bottom_border_color or right_border_color:
                existing_border = copy(cell.border)
                existing_bottom = copy(existing_border.bottom)
                existing_right = copy(existing_border.right)
                bottom_style = existing_bottom.style if existing_bottom is not None else None
                right_style = existing_right.style if existing_right is not None else None
                cell.border = Border(
                    left=existing_border.left,
                    right=(
                        Side(style=right_style or "medium", color=f"FF{right_border_color[1:].upper()}")
                        if right_border_color
                        else existing_border.right
                    ),
                    top=existing_border.top,
                    bottom=(
                        Side(style=bottom_style or "medium", color=f"FF{bottom_border_color[1:].upper()}")
                        if bottom_border_color
                        else existing_border.bottom
                    ),
                    diagonal=existing_border.diagonal,
                    diagonal_direction=existing_border.diagonal_direction,
                    outline=existing_border.outline,
                    vertical=existing_border.vertical,
                    horizontal=existing_border.horizontal,
                )

    # Save to the output buffer
    wb.save(output_buffer)
    output_buffer.seek(0)


def export_to_csv(df, output_buffer):
    """
    Export DataFrame to CSV with UTF-8 BOM for Excel compatibility.

    Args:
        output_buffer: BytesIO buffer to write to (use BytesIO for proper encoding handling)
    """
    # Write CSV to a StringIO first to get text, then encode with BOM
    temp_buffer = StringIO()
    df.to_csv(temp_buffer, index=False, header=False, lineterminator="\n")
    temp_buffer.seek(0)

    # Encode with UTF-8 BOM and write to output buffer
    csv_content = temp_buffer.getvalue()
    output_buffer.write(csv_content.encode("utf-8-sig"))
    output_buffer.seek(0)
